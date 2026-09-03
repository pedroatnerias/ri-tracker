"""Nucleo setorial auditavel dos indicadores operacionais de construcao civil."""

from __future__ import annotations

import math
import re
import unicodedata
from datetime import datetime, timezone
from typing import Any, Iterable

from company_registry import canonical_ticker, company_by_ticker
from construction_company_profiles import profile_for

EXTRACTOR_VERSION = "construction_operational_v1"
OWNERSHIP_BASES = {"company_share", "one_hundred_percent", "unknown"}
PERIOD_TYPES = {"quarter", "ytd_6m", "ytd_9m", "fy", "ltm", "ending_position"}
CALCULATION_STATES = {"found", "derived", "manual", "preserved_existing", "not_found", "incompatible_basis", "missing_components", "invalid_denominator", "extraction_error"}


def _metric(identifier: str, name: str, classification: str, definition: str, unit: str,
            nature: str, aliases: tuple[str, ...], positive: tuple[str, ...], negative: tuple[str, ...],
            bases: tuple[str, ...] = ("company_share", "one_hundred_percent", "unknown"),
            components: tuple[str, ...] = (), preferred_source: str = "official_operational_spreadsheet") -> dict[str, Any]:
    return {
        "id": identifier, "display_name": name, "sector": "construcao_civil",
        "classification": classification, "definition": definition, "unit": unit,
        "nature": nature, "expected_periodicity": "quarterly" if nature == "flow" else "quarter_end",
        "aliases": aliases, "positive_context": positive, "negative_context": negative,
        "accepted_bases": bases, "calculation_components": components,
        "validation_rules": ("explicit_period", "explicit_or_inferred_scale", "official_source", "no_double_count"),
        "comparability_rules": ("same_period_type", "same_currency", "same_scale", "same_ownership_basis", "same_segment"),
        "preferred_source": preferred_source,
        "methodology_notes": "Preserva total consolidado e aberturas; nunca soma o total com seus componentes.",
    }


CONSTRUCTION_OPERATIONAL_DICTIONARY: dict[str, dict[str, Any]] = {
    "landbank_vgv": _metric("landbank_vgv", "Banco de terras", "extracted", "VGV potencial do banco de terrenos ainda nao lancado.", "BRL_million", "stock", ("banco de terras", "landbank", "land bank", "estoque de terrenos", "vgv potencial", "potential sales value", "psv do landbank"), ("vgv", "psv", "potencial"), ("estoque de unidades", "valor contabil", "aquisicao de terrenos")),
    "launches_vgv": _metric("launches_vgv", "LanÃ§amentos", "extracted", "VGV de empreendimentos ou fases lancados no periodo.", "BRL_million", "flow", ("lancamentos", "vgv lancado", "launches", "launched psv", "volume lancado"), ("realizado", "periodo", "vgv"), ("guidance", "futuro", "unidades lancadas", "vendas")),
    "net_sales_vgv": _metric("net_sales_vgv", "VGV lÃ­quido â€” vendas lÃ­quidas", "extracted", "VGV de vendas apos distratos.", "BRL_million", "flow", ("vendas liquidas", "vgv liquido", "vendas liquidas contratadas", "net sales", "net contracted sales", "net psv sales"), ("contratadas", "vgv", "net"), ("receita liquida", "vendas brutas")),
    "cancellations_vgv": _metric("cancellations_vgv", "VGV distratado", "extracted", "VGV de contratos distratados no periodo.", "BRL_million", "flow", ("distratos", "vgv distratado", "cancelamentos", "cancellations", "terminated sales"), ("vgv", "periodo"), ("taxa de distrato", "provisao", "contingencia")),
    "units_sold": _metric("units_sold", "Lotes/unidades vendidas", "extracted", "Quantidade de lotes ou unidades vendidas no periodo.", "units", "flow", ("unidades vendidas", "lotes vendidos", "vendas em unidades", "unidades comercializadas", "units sold", "lots sold"), ("liquidas", "vendidas"), ("unidades lancadas", "entregues", "concluidas", "estoque")),
    "units_under_construction": _metric("units_under_construction", "Unidades em obras/em construÃ§Ã£o", "extracted", "Unidades com construcao iniciada e em andamento na data.", "units", "stock", ("unidades em obras", "unidades em construcao", "unidades em andamento", "unidades em producao", "units under construction", "units in progress"), ("unidades", "em obras"), ("numero de obras", "empreendimentos", "vgv", "unidades lancadas", "concluidas", "entregues")),
    "ending_inventory_vgv": _metric("ending_inventory_vgv", "Estoque EoP", "extracted", "VGV de unidades lancadas disponiveis ao fim do periodo.", "BRL_million", "stock", ("estoque eop", "estoque ao final do periodo", "estoque disponivel", "estoque a valor de mercado", "vgv em estoque", "inventory at market value", "ending inventory", "inventory psv"), ("vgv", "valor de mercado", "disponivel"), ("estoque contabil", "balanco patrimonial", "banco de terras", "terrenos")),
    "roe": _metric("roe", "ROE", "calculated", "Lucro liquido atribuivel LTM sobre PL atribuivel medio.", "percent", "flow", ("roe", "return on equity"), (), (), components=("controller_net_income_ltm", "controller_equity_begin", "controller_equity_end"), preferred_source="CVM_standardized_statements"),
    "credit_loss_allowance_to_receivables": _metric("credit_loss_allowance_to_receivables", "PCLD/RecebÃ­veis â€” proxy de inadimplÃªncia", "calculated", "Proxy de perdas esperadas sobre recebiveis brutos.", "percent", "stock", ("pcld sobre recebiveis", "allowance to receivables", "expected credit loss"), ("saldo", "contas a receber"), ("despesa", "dre", "contingencia"), components=("pcld_balance", "receivables_gross", "receivables_net"), preferred_source="CVM_balance_and_notes"),
    "net_vso": _metric("net_vso", "VSO lÃ­quida", "calculated", "Vendas liquidas sobre estoque inicial mais lancamentos.", "percent", "flow", ("vso liquida", "net vso", "sales over supply"), (), (), components=("net_sales_vgv", "beginning_inventory_vgv", "launches_vgv", "ending_inventory_vgv")),
}

CONSTRUCTION_METRIC_IDS = tuple(CONSTRUCTION_OPERATIONAL_DICTIONARY)


def repair_mojibake(value: Any) -> str:
    """Repairs the common UTF-8-read-as-Windows-1252 corruption.

    The conversion is applied only when typical mojibake markers exist and
    produces fewer markers, so normal Portuguese text is left untouched.
    """
    text = str(value or "")
    markers = ("\u00c3", "\u00c2", "\u00e2", "\u00f0", "\ufffd")
    if not any(marker in text for marker in markers):
        return text
    try:
        repaired = text.encode("cp1252").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return text
    marker_count = sum(text.count(marker) for marker in markers)
    repaired_count = sum(repaired.count(marker) for marker in markers)
    return repaired if repaired_count < marker_count else text


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", repair_mojibake(value))
    return " ".join("".join(ch for ch in text if not unicodedata.combining(ch)).lower().split())


def normalize_period(period: str) -> tuple[str, str]:
    text = normalize_text(period).replace(" ", "").upper().replace("Q", "T")
    match = re.fullmatch(r"([1-4])T(\d{2}|20\d{2})", text)
    if match:
        return f"{match.group(1)}T{match.group(2)[-2:]}", "quarter"
    match = re.fullmatch(r"([69])M(\d{2}|20\d{2})", text)
    if match:
        return f"{match.group(1)}M{match.group(2)[-2:]}", "ytd_6m" if match.group(1) == "6" else "ytd_9m"
    if re.fullmatch(r"(?:FY)?20\d{2}", text):
        return text[-4:], "fy"
    if re.fullmatch(r"FY\d{2}", text):
        return f"20{text[-2:]}", "fy"
    if text.startswith("LTM"):
        return text, "ltm"
    raise ValueError(f"periodo operacional invalido: {period}")


def infer_ownership_basis(text: str) -> str:
    value = normalize_text(text)
    if any(term in value for term in ("participacao da companhia", "company share", "% companhia", "%avll", "% cal", "%ez", "% even", "%cbr", "ex permuta")):
        return "company_share"
    if any(term in value for term in ("100%", "cem por cento", "total projects")):
        return "one_hundred_percent"
    return "unknown"


def normalize_money(value: float, unit_text: str) -> tuple[float, str, str]:
    unit = normalize_text(unit_text)
    number = float(value)
    if "bilh" in unit or "billion" in unit:
        return number * 1000, "BRL", "million"
    if re.search(r"\bmil\b|thousand", unit) and "milhao" not in unit and "million" not in unit:
        return number / 1000, "BRL", "million"
    if re.search(r"(?:r\$|brl)\s*mm\b", unit) or "milhoes" in unit or "milhao" in unit or "million" in unit:
        return number, "BRL", "million"
    if "r$" in unit and not any(term in unit for term in ("milhao", "million", " mm")):
        return number / 1_000_000, "BRL", "million"
    return number, "BRL", "million"


def resolve_financial_unit(unit_text: str, source: str = "column_header") -> dict[str, Any]:
    unit = normalize_text(unit_text)
    if "bilh" in unit or "billion" in unit:
        multiplier = 1000
    elif re.search(r"\bmil\b|thousand", unit) and "milhao" not in unit and "million" not in unit:
        multiplier = .001
    elif re.search(r"(?:r\$|brl)\s*mm\b", unit) or "milhoes" in unit or "milhao" in unit or "million" in unit:
        multiplier = 1
    elif "r$" in unit or "brl" in unit:
        multiplier = .000001
    else:
        multiplier = 1
        source = "unknown"
    return {
        "currency": "BRL",
        "raw_scale": unit_text,
        "normalized_unit": "BRL_million",
        "multiplier": multiplier,
        "source": source,
        "confidence": "high" if source != "unknown" else "low",
    }


def parse_brazilian_financial_value(raw_value: str | int | float, declared_scale: str) -> dict[str, Any]:
    raw = str(raw_value).strip()
    if isinstance(raw_value, (int, float)):
        parsed = float(raw_value)
    else:
        cleaned = re.sub(r"[^\d,.-]", "", raw)
        if not cleaned:
            raise ValueError(f"valor financeiro invÃ¡lido: {raw_value}")
        if "," in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        elif re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+", cleaned):
            cleaned = cleaned.replace(".", "")
        parsed = float(cleaned)
    unit = resolve_financial_unit(declared_scale)
    normalized = parsed * float(unit["multiplier"])
    return {
        "raw_value": raw, "raw_unit": declared_scale, "parsed_value": parsed,
        "normalized_value": normalized, "normalized_unit": "BRL_million",
        "currency": unit["currency"], "scale": "million",
        "multiplier": unit["multiplier"], "unit_source": unit["source"],
        "conversion_steps": [{"from": raw, "parsed_value": parsed}, {"multiplier": unit["multiplier"], "normalized_unit": "BRL_million"}],
        "scale_conversion_applied": not math.isclose(parsed, normalized),
    }


def parse_composite_header(cell: str) -> dict[str, Any]:
    parts = [part.strip() for part in re.split(r"<br\s*/?>|\n", str(cell), flags=re.I) if part.strip()]
    title = parts[0] if parts else ""
    periods: list[str] = []
    for part in parts[1:] or parts:
        for token in re.findall(r"(?:\b(?:[1-4]T\d{2,4}|[69]M\d{2,4}|FY\d{2,4}|20\d{2})\b|Var%)", part, flags=re.I):
            periods.append(token.upper())
    return {"title": title, "periods": periods, "unit": title}


def split_composite_cell(cell: str) -> list[str]:
    return [part.strip() for part in re.split(r"<br\s*/?>|\n", str(cell), flags=re.I) if part.strip()]


def align_periods_and_values(headers: list[str], values: list[str]) -> list[tuple[str, str]]:
    expanded_headers: list[str] = []
    for header in headers:
        parsed = parse_composite_header(header)
        expanded_headers.extend(parsed["periods"] or [header])
    expanded_values: list[str] = []
    for value in values:
        expanded_values.extend(split_composite_cell(value))
    # A positional association is unsafe when the table has been truncated or
    # a cell contains an unlabelled extra value. Reject it instead of silently
    # pairing values with the wrong period.
    if len(expanded_headers) != len(expanded_values):
        return []
    aligned: list[tuple[str, str]] = []
    for period, value in zip(expanded_headers, expanded_values):
        if normalize_text(period) in {"var", "var%"} or "%" in period and not re.search(r"\d", period):
            continue
        if "%" in str(value):
            continue
        aligned.append((period, value))
    return aligned


def validate_observation_evidence(observation: dict[str, Any]) -> dict[str, Any]:
    missing = [key for key in ("ticker", "indicator_id", "period", "value", "unit", "source_document") if not observation.get(key)]
    status = "valid" if not missing and observation.get("confidence") in {"high", "medium"} else "low_confidence"
    if observation.get("validation_flags"):
        if "breakdown_without_explicit_total" in observation["validation_flags"]:
            status = "quarantined_scope"
        elif any(flag.startswith(("profile_publication:", "ownership_basis_expected:", "unexpected_")) for flag in observation["validation_flags"]):
            status = "quarantined"
    return {**observation, "validation_status": status, "validation_missing_fields": missing}


def extract_table_observations(table: list[list[str]], context: dict[str, Any]) -> list[dict[str, Any]]:
    if not table:
        return []
    headers = table[0]
    observations: list[dict[str, Any]] = []
    for row in table[1:]:
        if not row:
            continue
        label = row[0]
        context_text = " ".join((str(context.get("ticker") or ""), str(context.get("table_title") or ""), " ".join(headers)))
        indicator_id, flags = identify_metric(label, context_text, " ".join(headers))
        if not indicator_id or flags:
            continue
        for period, raw_value in align_periods_and_values(headers[1:], row[1:]):
            try:
                parsed = parse_brazilian_financial_value(raw_value, " ".join(headers))
                observation = build_evidence_observation(
                    ticker=str(context.get("ticker") or ""),
                    indicator_id=indicator_id,
                    value=parsed["parsed_value"],
                    period=period,
                    label=label,
                    unit=" ".join(headers),
                    context=context_text,
                    source_document=str(context.get("source_document") or ""),
                    source_url=str(context.get("source_url") or ""),
                    table_title=context_text,
                    column_label=period,
                    raw_value=raw_value,
                    raw_unit=" ".join(headers),
                )
            except (ValueError, KeyError):
                continue
            observations.append(validate_observation_evidence(observation))
    return observations


def _markdown_page_for_line(lines: list[str], index: int) -> int | None:
    for line in reversed(lines[: index + 1]):
        match = re.search(r"(?:pagina|page)\s*(\d+)", normalize_text(line))
        if match:
            return int(match.group(1))
    return None


def _markdown_tables(lines: list[str]) -> list[tuple[list[list[str]], int, str]]:
    tables: list[tuple[list[list[str]], int, str]] = []
    current: list[list[str]] = []
    start = 0
    title = ""
    for index, raw_line in enumerate(lines):
        line = raw_line.strip()
        if line.startswith("#"):
            title = line.lstrip("# ").strip()
        if line.startswith("|") and line.endswith("|"):
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if all(re.fullmatch(r":?-+:?", cell) for cell in cells):
                continue
            if not current:
                start = index
            current.append(cells)
            continue
        if current:
            tables.append((current, start, title))
            current = []
    if current:
        tables.append((current, start, title))
    return tables


def _dedupe_observations(observations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique: dict[tuple[Any, ...], dict[str, Any]] = {}
    for observation in observations:
        key = (observation["ticker"], observation["indicator_id"], observation["period"], observation["ownership_basis"], observation["segment"])
        current = unique.get(key)
        score = (
            observation["consolidated_or_breakdown"] == "consolidated",
            observation["ownership_basis"] == "company_share",
            {"high": 2, "medium": 1}.get(observation["confidence"], 0),
            observation.get("source_type") == "official_spreadsheet",
        )
        current_score = (-1, -1, -1, -1) if current is None else (
            current["consolidated_or_breakdown"] == "consolidated",
            current["ownership_basis"] == "company_share",
            {"high": 2, "medium": 1}.get(current["confidence"], 0),
            current.get("source_type") == "official_spreadsheet",
        )
        if current is None or score > current_score:
            unique[key] = observation
    return list(unique.values())


def identify_metric(label: str, context: str = "", unit: str = "") -> tuple[str | None, list[str]]:
    haystack = normalize_text(" ".join((label, context)))
    matches: list[tuple[int, str]] = []
    flags: list[str] = []
    for metric_id, definition in CONSTRUCTION_OPERATIONAL_DICTIONARY.items():
        if definition["classification"] != "extracted":
            continue
        profile = profile_for(_ticker_from_context(context))
        rules = profile.get("metrics", {}).get(metric_id, {})
        aliases = tuple(definition["aliases"]) + tuple(rules.get("positive_aliases", []))
        forbidden = tuple(definition["negative_context"]) + tuple(rules.get("negative_aliases", []))
        alias_hits = [a for a in aliases if normalize_text(a) in haystack]
        if not alias_hits:
            continue
        negatives = [n for n in forbidden if normalize_text(n) in haystack]
        if negatives:
            flags.extend(f"negative_context:{normalize_text(n)}" for n in negatives)
            continue
        score = max(len(normalize_text(a)) for a in alias_hits)
        if metric_id.endswith("_vgv") and not any(token in normalize_text(unit + " " + haystack) for token in ("r$", "brl", "vgv", "psv", "valor de mercado")):
            flags.append("missing_monetary_context")
            continue
        matches.append((score, metric_id))
    matches.sort(reverse=True)
    return (matches[0][1] if matches else None), flags


def _ticker_from_context(context: str) -> str:
    """Resolve an optional ticker embedded in extraction context.

    Keeping this lookup tolerant preserves the generic parser for documents
    without metadata while allowing declarative company rules when available.
    """
    match = re.search(r"\b([A-Z]{4,5}3)\b", str(context or "").upper())
    return match.group(1) if match else ""


def build_evidence_observation(*, ticker: str, indicator_id: str, value: float, period: str,
        label: str, unit: str, context: str = "", source_document: str = "", source_url: str = "",
        source_type: str = "release", page: int | None = None, table_title: str = "",
        column_label: str = "", ownership_basis: str | None = None, segment: str = "consolidated",
        region: str = "", reported_or_derived: str = "reported", raw_value: str | None = None,
        raw_unit: str | None = None) -> dict[str, Any]:
    definition = CONSTRUCTION_OPERATIONAL_DICTIONARY[indicator_id]
    canonical_period, period_type = normalize_period(period)
    if definition["nature"] == "stock":
        period_type = "ending_position"
    basis = ownership_basis or infer_ownership_basis(" ".join((label, context, table_title)))
    if basis not in OWNERSHIP_BASES:
        raise ValueError(f"ownership_basis invalida: {basis}")
    normalized_value, currency, scale = (normalize_money(value, unit) if indicator_id.endswith("_vgv") else (float(value), None, "units"))
    rules = profile_for(ticker).get("metrics", {}).get(indicator_id, {})
    validation_flags: list[str] = []
    expected_basis = rules.get("preferred_ownership_basis")
    if expected_basis and basis != expected_basis:
        validation_flags.append(f"ownership_basis_expected:{expected_basis}")
    expected_sign = rules.get("expected_sign")
    if expected_sign == "negative" and normalized_value > 0:
        validation_flags.append("unexpected_positive_sign")
    if expected_sign == "positive" and normalized_value < 0:
        validation_flags.append("unexpected_negative_sign")
    if rules.get("publication") in {"ambiguous", "not_disclosed", "not_applicable"}:
        validation_flags.append(f"profile_publication:{rules['publication']}")
    confidence_reasons = ["official_source" if source_type not in {"manual", "secondary"} else source_type, "indicator_explicit", "period_confirmed"]
    confidence = "high" if basis != "unknown" and unit else "medium"
    return {
        "sector": "construcao_civil", "ticker": canonical_ticker(ticker), "ticker_found": ticker.upper(),
        "company": company_by_ticker(ticker).expected_name, "indicator_id": indicator_id,
        "indicator_name": definition["display_name"], "value": normalized_value, "unit": definition["unit"],
        "currency": currency, "scale": scale, "period": canonical_period, "period_type": period_type,
        "reference_date": None, "flow_or_stock": definition["nature"], "ownership_basis": basis,
        "gross_or_net": "net" if indicator_id in {"net_sales_vgv", "net_vso"} else "not_applicable",
        "segment": segment, "region": region, "consolidated_or_breakdown": "consolidated" if segment == "consolidated" else "breakdown",
        "source_document": source_document, "source_url": source_url, "source_type": source_type,
        "page": page, "table_title": table_title, "row_label": label, "column_label": column_label or period,
        "evidence_text": context or label, "extraction_method": "contextual_table",
        "reported_or_derived": reported_or_derived, "confidence": confidence,
        "confidence_reasons": confidence_reasons, "validation_flags": validation_flags + ([] if basis != "unknown" else ["unknown_ownership_basis"]),
        "raw_value": raw_value if raw_value is not None else str(value), "raw_unit": raw_unit or unit,
        "normalized_value": normalized_value, "normalized_unit": definition["unit"],
        "scale_conversion_applied": not math.isclose(float(value), normalized_value) if indicator_id.endswith("_vgv") else False,
        "created_at": datetime.now(timezone.utc).isoformat(), "extractor_version": EXTRACTOR_VERSION,
    }


def extract_markdown_observations(text: str, *, ticker: str, source_document: str = "",
                                  source_url: str = "") -> list[dict[str, Any]]:
    """Extrai linhas de tabelas Markdown; exige rotulo, periodo e valor explicitos."""
    lines = text.splitlines()
    observations: list[dict[str, Any]] = []
    for table, start, title in _markdown_tables(lines):
        rows = extract_table_observations(
            table,
            {"ticker": ticker, "source_document": source_document, "source_url": source_url, "table_title": title},
        )
        page = _markdown_page_for_line(lines, start)
        for row in rows:
            if page and not row.get("page"):
                row["page"] = page
            row["extraction_method"] = "markdown_composite_table"
            observations.append(row)
    table_title = ""
    page: int | None = None
    headers: list[str] = []
    for index, raw_line in enumerate(lines):
        line = raw_line.strip()
        page_match = re.search(r"(?:pagina|page)\s*(\d+)", normalize_text(line))
        if page_match:
            page = int(page_match.group(1))
        if line.startswith("#"):
            table_title = line.lstrip("# ").strip()
        if not (line.startswith("|") and line.endswith("|")):
            continue
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if len(cells) < 2 or all(re.fullmatch(r":?-+:?", cell) for cell in cells):
            continue
        if any(re.fullmatch(r"(?:[1-4][TQ]\s?\d{2,4}|[69]M\s?\d{2,4}|FY\s?\d{2,4}|20\d{2})", cell, re.I) for cell in cells[1:]):
            headers = cells
            continue
        if not headers or len(cells) < 2:
            continue
        context = " ".join((ticker, *lines[max(0, index - 3): min(len(lines), index + 4)]))
        indicator_id, flags = identify_metric(cells[0], context, context)
        normalized_context = normalize_text(context)
        normalized_label = normalize_text(cells[0])
        if any(term in normalized_context for term in ("por regiao", "por produto", "by region", "by product")) and not any(term in normalized_label for term in ("total", "consolidado", "consolidated")):
            flags.append("breakdown_without_explicit_total")
        if not indicator_id or flags:
            continue
        for column, raw_value in enumerate(cells[1:], start=1):
            if column >= len(headers) or raw_value in {"", "-", "â€”", "N/A"}:
                continue
            cleaned = re.sub(r"[^\d,.-]", "", raw_value)
            if not cleaned:
                continue
            if "," in cleaned:
                cleaned = cleaned.replace(".", "").replace(",", ".")
            elif re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+", cleaned):
                cleaned = cleaned.replace(".", "")
            try:
                value = float(cleaned)
                observation = build_evidence_observation(
                    ticker=ticker, indicator_id=indicator_id, value=value, period=headers[column],
                    label=cells[0], unit=" ".join((cells[0], table_title, context)), context=context,
                    source_document=source_document, source_url=source_url, page=page,
                    table_title=table_title, column_label=headers[column],
                    raw_value=raw_value, raw_unit=" ".join((cells[0], table_title, context)),
                )
            except (ValueError, KeyError):
                continue
            observations.append(observation)
    return _dedupe_observations(observations)


def extract_workbook_observations(workbook_path: Any, *, ticker: str, source_document: str = "", source_url: str = "") -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    observations: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            merged_values: dict[str, Any] = {}
            for merged in sheet.merged_cells.ranges:
                value = sheet.cell(merged.min_row, merged.min_col).value
                for row in range(merged.min_row, merged.max_row + 1):
                    for column in range(merged.min_col, merged.max_col + 1):
                        merged_values[sheet.cell(row, column).coordinate] = value
            rows: list[list[str]] = []
            cells_by_row: list[list[str]] = []
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 0, 300), min_col=1, max_col=min(sheet.max_column or 0, 80)):
                values: list[str] = []
                refs: list[str] = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        value = merged_values.get(cell.coordinate)
                    values.append("" if value is None else str(value))
                    refs.append(cell.coordinate)
                if any(value.strip() for value in values):
                    rows.append(values)
                    cells_by_row.append(refs)
            for index in range(max(0, len(rows) - 1)):
                header = rows[index]
                data = rows[index + 1]
                if not any(parse_composite_header(cell)["periods"] for cell in header):
                    continue
                context = {
                    "ticker": ticker,
                    "source_document": source_document or str(workbook_path),
                    "source_url": source_url,
                    "table_title": sheet.title,
                }
                extracted = extract_table_observations([header, data], context)
                for observation in extracted:
                    observation["source_type"] = "official_spreadsheet"
                    observation["sheet"] = sheet.title
                    observation["source_cell"] = ",".join(cells_by_row[index + 1][: len(data)])
                    observations.append(observation)
    finally:
        workbook.close()
    return _dedupe_observations(observations)


def calculate_roe(net_income_ltm: float | None, equity_begin: float | None, equity_end: float | None,
                  *, scopes_match: bool = True) -> dict[str, Any]:
    components = {"controller_net_income_ltm": net_income_ltm, "controller_equity_begin": equity_begin, "controller_equity_end": equity_end}
    if None in components.values():
        return {"indicator_id": "roe", "value": None, "calculation_status": "missing_components", "components": components}
    if not scopes_match:
        return {"indicator_id": "roe", "value": None, "calculation_status": "incompatible_basis", "components": components}
    average = (float(equity_begin) + float(equity_end)) / 2
    if average <= 0:
        return {"indicator_id": "roe", "value": None, "calculation_status": "invalid_denominator", "invalid_denominator": True, "average_equity": average, "components": components}
    return {"indicator_id": "roe", "value": float(net_income_ltm) / average, "unit": "percent", "calculation_status": "derived", "formula": "controller_net_income_ltm / average_controller_equity", "average_equity": average, "components": components}


def calculate_credit_loss_proxy(pcld_balance: float | None, *, receivables_gross: float | None = None,
                                receivables_net: float | None = None, same_scope: bool = True) -> dict[str, Any]:
    allowance = None if pcld_balance is None else abs(float(pcld_balance))
    reconstructed = False
    gross = None if receivables_gross is None else float(receivables_gross)
    if gross is None and receivables_net is not None and allowance is not None:
        gross, reconstructed = float(receivables_net) + allowance, True
    payload = {"indicator_id": "credit_loss_allowance_to_receivables", "pcld_balance": allowance,
               "receivables_net": receivables_net, "receivables_gross": gross,
               "gross_receivables_reconstructed": reconstructed,
               "methodology_warning": "PCLD/RecebÃ­veis Ã© uma proxy baseada em perdas esperadas e pode refletir tanto risco de crÃ©dito quanto diferenÃ§as nas polÃ­ticas de provisionamento."}
    if allowance is None or gross is None:
        return {**payload, "value": None, "calculation_status": "missing_components"}
    if not same_scope:
        return {**payload, "value": None, "calculation_status": "incompatible_basis"}
    if gross <= 0:
        return {**payload, "value": None, "calculation_status": "invalid_denominator", "invalid_denominator": True}
    return {**payload, "value": allowance / gross, "proxy_value": allowance / gross, "unit": "percent", "calculation_status": "derived"}


def calculate_vso(net_sales_vgv: float | None, beginning_inventory_vgv: float | None,
                  launches_vgv: float | None, *, ending_inventory_vgv: float | None = None,
                  ownership_bases: Iterable[str] = (), compatible_periods: bool = True,
                  reported_vso: float | None = None, materiality: float = .05) -> dict[str, Any]:
    components = {"net_sales_vgv": net_sales_vgv, "beginning_inventory_vgv": beginning_inventory_vgv,
                  "launches_vgv": launches_vgv, "ending_inventory_vgv": ending_inventory_vgv}
    missing = [key for key in ("net_sales_vgv", "beginning_inventory_vgv", "launches_vgv") if components[key] is None]
    bases = {basis for basis in ownership_bases if basis}
    base_incompatible = len(bases) > 1 or "unknown" in bases
    result = {"indicator_id": "net_vso", **components, "reported_vso": reported_vso,
              "missing_components": missing, "ownership_basis": next(iter(bases)) if len(bases) == 1 else "unknown"}
    if missing:
        return {**result, "value": None, "calculated_vso": None, "calculation_status": "missing_components", "reconciliation_status": "not_calculable"}
    if base_incompatible or not compatible_periods:
        return {**result, "value": None, "calculated_vso": None, "calculation_status": "incompatible_basis", "reconciliation_status": "not_calculable"}
    denominator = float(beginning_inventory_vgv) + float(launches_vgv)
    if denominator <= 0:
        return {**result, "value": None, "calculated_vso": None, "calculation_status": "invalid_denominator", "reconciliation_status": "not_calculable"}
    value = float(net_sales_vgv) / denominator
    adjustment = None if ending_inventory_vgv is None else float(ending_inventory_vgv) - (denominator - float(net_sales_vgv))
    if adjustment is None:
        reconciliation = "not_calculable"
    elif math.isclose(adjustment, 0, abs_tol=1e-9):
        reconciliation = "reconciled"
    elif abs(adjustment) / max(abs(denominator), 1) <= materiality:
        reconciliation = "immaterial_difference"
    else:
        reconciliation = "material_difference"
    return {**result, "value": value, "calculated_vso": value, "unit": "percent", "calculation_status": "derived",
            "implicit_adjustments": adjustment, "reconciliation_status": reconciliation,
            "reported_calculated_difference": None if reported_vso is None else value - float(reported_vso),
            "confidence": "medium" if reconciliation == "material_difference" else "high"}


def calculate_derived_from_observations(observations: list[dict[str, Any]]) -> dict[str, Any]:
    """Calculate only metrics with compatible components in the same snapshot."""
    by_key = {(item.get("indicator_id"), item.get("period"), item.get("ownership_basis")): item.get("value") for item in observations}
    vso: dict[str, Any] = {}
    periods = {item.get("period") for item in observations if item.get("period")}
    def period_key(value: str) -> tuple[int, int] | None:
        match = re.fullmatch(r"([1-4])T(\d{2})", str(value))
        return (2000 + int(match.group(2)), int(match.group(1))) if match else None
    for period in periods:
        bases = {item.get("ownership_basis") for item in observations if item.get("period") == period}
        for basis in bases - {None, "unknown"}:
            key = period_key(str(period))
            previous = None
            if key:
                previous_periods = [candidate for candidate in periods if period_key(str(candidate)) and period_key(str(candidate)) < key]
                if previous_periods:
                    previous = max(previous_periods, key=lambda candidate: period_key(str(candidate)))
            beginning_inventory = by_key.get(("ending_inventory_vgv", previous, basis)) if previous else None
            result = calculate_vso(by_key.get(("net_sales_vgv", period, basis)), beginning_inventory, by_key.get(("launches_vgv", period, basis)), ownership_bases=(basis,), compatible_periods=beginning_inventory is not None)
            if result["calculation_status"] == "derived":
                vso[period] = result
    return {"roe": {"value": None, "calculation_status": "missing_components"}, "credit_loss_allowance_to_receivables": {"value": None, "calculation_status": "missing_components"}, "net_vso": vso or {"value": None, "calculation_status": "missing_components"}}
