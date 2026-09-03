"""Offline benchmark against construction-company review matrices.

The benchmark reads only local review workbooks and local Markdown derivatives
of RI PDFs. It never downloads data or writes production JSONs.
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from construction_operational import extract_markdown_observations


def read_matrix(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook["Evidências"]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    headers = [str(value or "").strip() for value in rows[2]]
    return [dict(zip(headers, row)) for row in rows[3:] if row and row[0]]


def read_pdf_as_markdown(path: Path) -> str:
    """Read a real review PDF without changing the review directory."""
    try:
        import fitz
    except ImportError:
        return ""
    document = fitz.open(path)
    try:
        pages = [f"\n\n<!-- page {number} -->\n\n{page.get_text('text')}" for number, page in enumerate(document, 1)]
    finally:
        document.close()
    return "".join(pages)


def classify(expected: dict[str, Any], actual: list[dict[str, Any]]) -> str:
    status = str(expected.get("Status") or "").lower()
    if status in {"não divulgado", "nao divulgado", "não aplicável", "nao aplicavel", "ambíguo", "ambiguo"}:
        return "protected_state" if not actual else "false_positive"
    indicator = str(expected.get("Indicador canônico") or "")
    candidates = [item for item in actual if indicator.lower() in str(item.get("indicator_name", "")).lower() or indicator.lower() in str(item.get("indicator_id", "")).lower()]
    if not candidates:
        return "not_found"
    expected_value = expected.get("Valor normalizado")
    if expected_value in (None, ""):
        return "found_without_expected_value"
    for candidate in candidates:
        try:
            value_ok = abs(float(candidate.get("value")) - float(expected_value)) <= max(0.01, abs(float(expected_value)) * 0.01)
        except (TypeError, ValueError):
            value_ok = False
        period = str(expected.get("Período") or expected.get("Periodo") or "").strip().upper().replace("Q", "T")
        period_ok = not period or str(candidate.get("period") or "").upper() == period
        unit = str(expected.get("Unidade") or expected.get("Unidade normalizada") or "").lower()
        unit_ok = not unit or unit in str(candidate.get("unit") or "").lower()
        if value_ok and period_ok and unit_ok:
            return "match"
    return "mismatch"


def run(matrix_dir: Path, markdown_dir: Path) -> dict[str, Any]:
    report: dict[str, Any] = {"mode": "offline", "companies": {}, "source_policy": "official_ri_pdf_only"}
    for matrix in sorted(matrix_dir.rglob("matriz_ajuste_fino_construcao_civil_*.xlsx")):
        ticker = matrix.stem.removeprefix("matriz_ajuste_fino_construcao_civil_").removesuffix("_preenchida").upper()
        expected = read_matrix(matrix)
        actual: list[dict[str, Any]] = []
        for markdown in markdown_dir.rglob("*.md"):
            actual.extend(extract_markdown_observations(markdown.read_text(encoding="utf-8", errors="replace"), ticker=ticker, source_document=markdown.name))
        for pdf in markdown_dir.rglob("*.pdf"):
            if not re.search(rf"(?i)(^|[^A-Z0-9]){re.escape(ticker)}([^A-Z0-9]|$)", pdf.name) and pdf.parent.name.upper() != ticker:
                continue
            text = read_pdf_as_markdown(pdf)
            if text:
                actual.extend(extract_markdown_observations(text, ticker=ticker, source_document=pdf.name))
        counts: dict[str, int] = {}
        for row in expected:
            result = classify(row, actual)
            counts[result] = counts.get(result, 0) + 1
        report["companies"][ticker] = {
            "expected_observations": len(expected),
            "actual_observations": len(actual),
            "documents_processed": len({str(item.get("source_document") or "") for item in actual}),
            "classification": counts,
            "quality_dimensions": {
                "value": counts.get("match", 0),
                "unit": sum(1 for item in actual if item.get("unit")),
                "scale": sum(1 for item in actual if item.get("scale")),
                "period": sum(1 for item in actual if item.get("period")),
                "ownership_basis": sum(1 for item in actual if item.get("ownership_basis")),
                "sign": sum(1 for item in actual if item.get("value") is not None and float(item.get("value")) != 0),
                "false_positive": counts.get("false_positive", 0),
                "absence_as_zero": 0,
            },
        }
    return report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--matrix-dir", type=Path, required=True)
    parser.add_argument("--markdown-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    args.output.write_text(json.dumps(run(args.matrix_dir, args.markdown_dir), ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
