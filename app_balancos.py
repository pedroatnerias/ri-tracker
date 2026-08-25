#!/usr/bin/env python3
"""Descarrega ITRs da CVM e monta balanços patrimoniais em Excel."""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import shutil
import sys
import tempfile
import time
import unicodedata
import urllib.error
import urllib.request
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from company_registry import financial_companies
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_URL = (
    "https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/ITR/DADOS/"
    "itr_cia_aberta_{year}.zip"
)
DFP_BASE_URL = (
    "https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/DFP/DADOS/"
    "dfp_cia_aberta_{year}.zip"
)
SOURCE_PAGE = "https://dados.cvm.gov.br/dataset/cia_aberta-doc-itr"
DFP_SOURCE_PAGE = "https://dados.cvm.gov.br/dataset/cia_aberta-doc-dfp"
USER_AGENT = "Mozilla/5.0 (compatible; CVM-ITR-Balancos/1.0)"


@dataclass(frozen=True)
class Company:
    ticker: str
    cd_cvm: str
    cnpj: str
    statement_scope: str  # "con" ou "ind"
    expected_name: str

    @property
    def scope_label(self) -> str:
        return "Consolidado" if self.statement_scope == "con" else "Individual"


# O código CVM é a chave principal. O CNPJ funciona como validação/fallback.
COMPANIES = (
    Company("AALR3", "024023", "42.771.949/0001-35", "con", "CENTRO DE IMAGEM DIAGNOSTICOS S.A."),
    Company("DASA3", "019623", "61.486.650/0001-83", "con", "DIAGNOSTICOS DA AMERICA S.A."),
    Company("FLRY3", "021881", "60.840.055/0001-31", "con", "FLEURY S.A."),
    Company("HAPV3", "024392", "05.197.443/0001-38", "con", "HAPVIDA PARTICIPACOES E INVESTIMENTOS S.A."),
    Company("MATD3", "025690", "16.676.520/0001-59", "con", "HOSPITAL MATER DEI S.A."),
    Company("ONCO3", "026123", "12.104.241/0004-02", "con", "ONCOCLINICAS DO BRASIL SERVICOS MEDICOS S.A."),
    Company("RDOR3", "024821", "06.047.087/0001-39", "ind", "REDE D'OR SAO LUIZ S.A."),
)

READ_COLUMNS = [
    "CNPJ_CIA",
    "DT_REFER",
    "VERSAO",
    "DENOM_CIA",
    "CD_CVM",
    "MOEDA",
    "ESCALA_MOEDA",
    "ORDEM_EXERC",
    "DT_FIM_EXERC",
    "CD_CONTA",
    "DS_CONTA",
    "VL_CONTA",
    "ST_CONTA_FIXA",
]

FATORES_ESCALA = {
    "UNIDADE": 1,
    "UNIDADES": 1,
    "MIL": 1_000,
    "MILHAR": 1_000,
    "MILHARES": 1_000,
    "MILHAO": 1_000_000,
    "MILHOES": 1_000_000,
}

NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREY = "E7E6E6"
GREEN = "E2F0D9"
RED = "FCE4D6"
WHITE = "FFFFFF"
BLACK = "000000"
THIN_GREY = Side(style="thin", color="B7B7B7")


def parse_args() -> argparse.Namespace:
    current_year = date.today().year
    parser = argparse.ArgumentParser(
        description="Descarrega ITRs e monta balanços patrimoniais no padrão CVM."
    )
    parser.add_argument(
        "--years",
        nargs="+",
        type=int,
        default=list(range(current_year - 4, current_year + 1)),
        help="Anos dos ITRs. Padrão: ano corrente e quatro anteriores.",
    )
    parser.add_argument(
        "--output-dir", type=Path, default=Path("resultados"), help="Pasta de saída."
    )
    parser.add_argument(
        "--offline",
        action="store_true",
        help="Não descarrega; exige todos os ZIPs na pasta downloads.",
    )
    parser.add_argument(
        "--force-download", action="store_true", help="Substitui ZIPs já existentes."
    )
    parser.add_argument("--no-dfp", action="store_true", help="Nao incorpora DFPs anuais.")
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--sector", choices=("saude", "construcao_civil", "all"), default="saude")
    args = parser.parse_args()
    args.years = sorted(set(args.years))
    if not args.years:
        parser.error("Informe ao menos um ano.")
    if min(args.years) < 2011 or max(args.years) > current_year:
        parser.error(f"Os anos devem estar entre 2011 e {current_year}.")
    return args


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c)).upper().strip()


def normalize_cd_cvm(series: pd.Series) -> pd.Series:
    return series.astype("string").str.replace(r"\.0$", "", regex=True).str.zfill(6)


def valid_zip(path: Path, year: int, doc: str = "itr") -> bool:
    if not path.exists() or path.stat().st_size == 0:
        return False
    prefix = "dfp" if doc == "dfp" else "itr"
    expected = {
        f"{prefix}_cia_aberta_BPA_con_{year}.csv",
        f"{prefix}_cia_aberta_BPP_con_{year}.csv",
        f"{prefix}_cia_aberta_BPA_ind_{year}.csv",
        f"{prefix}_cia_aberta_BPP_ind_{year}.csv",
    }
    try:
        with zipfile.ZipFile(path) as archive:
            return archive.testzip() is None and expected.issubset(set(archive.namelist()))
    except (OSError, zipfile.BadZipFile):
        return False


def download_zip(year: int, destination: Path, force: bool, offline: bool, doc: str = "itr") -> Path:
    destination.mkdir(parents=True, exist_ok=True)
    prefix = "dfp" if doc == "dfp" else "itr"
    output = destination / f"{prefix}_cia_aberta_{year}.zip"

    if valid_zip(output, year, doc) and not force:
        logging.info("ZIP %s já existe e é válido.", year)
        return output
    if offline:
        raise FileNotFoundError(f"ZIP válido de {year} não encontrado em {destination}.")

    url = (DFP_BASE_URL if doc == "dfp" else BASE_URL).format(year=year)
    logging.info("A descarregar %s", url)
    last_error: Exception | None = None

    for attempt in range(1, 4):
        temporary: Path | None = None
        try:
            request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(request, timeout=120) as response:
                content_type = response.headers.get("Content-Type", "").lower()
                if "html" in content_type:
                    raise RuntimeError(f"A CVM devolveu HTML em vez de ZIP para {year}.")
                with tempfile.NamedTemporaryFile(
                    prefix=f"itr_{year}_", suffix=".zip", dir=destination, delete=False
                ) as temp_file:
                    temporary = Path(temp_file.name)
                    shutil.copyfileobj(response, temp_file, length=1024 * 1024)
            if not valid_zip(temporary, year, doc):
                raise zipfile.BadZipFile(f"Download de {year} incompleto ou inválido.")
            os.replace(temporary, output)
            return output
        except (urllib.error.URLError, TimeoutError, OSError, RuntimeError, zipfile.BadZipFile) as exc:
            last_error = exc
            if temporary and temporary.exists():
                temporary.unlink()
            if attempt < 3:
                wait = 2**attempt
                logging.warning("Falha na tentativa %d/3 (%s). Nova tentativa em %ds.", attempt, exc, wait)
                time.sleep(wait)

    raise RuntimeError(f"Não foi possível descarregar o ITR de {year}: {last_error}")


def read_company_statement(
    archive_path: Path, year: int, company: Company, statement: str, doc: str = "itr"
) -> pd.DataFrame:
    prefix = "dfp" if doc == "dfp" else "itr"
    member = f"{prefix}_cia_aberta_{statement}_{company.statement_scope}_{year}.csv"
    pieces: list[pd.DataFrame] = []

    with zipfile.ZipFile(archive_path) as archive:
        if member not in archive.namelist():
            raise KeyError(f"{member} não existe em {archive_path.name}.")
        with archive.open(member) as csv_file:
            chunks = pd.read_csv(
                csv_file,
                sep=";",
                encoding="latin1",
                decimal=".",
                usecols=READ_COLUMNS,
                dtype={
                    "CNPJ_CIA": "string",
                    "DENOM_CIA": "string",
                    "CD_CVM": "string",
                    "CD_CONTA": "string",
                    "DS_CONTA": "string",
                    "ST_CONTA_FIXA": "string",
                    "MOEDA": "string",
                    "ESCALA_MOEDA": "string",
                    "ORDEM_EXERC": "string",
                },
                chunksize=100_000,
                low_memory=False,
            )
            for chunk in chunks:
                chunk["CD_CVM"] = normalize_cd_cvm(chunk["CD_CVM"])
                by_code = chunk["CD_CVM"].eq(company.cd_cvm)
                by_cnpj = chunk["CNPJ_CIA"].eq(company.cnpj)
                selected = chunk[by_code | by_cnpj].copy()
                if not selected.empty:
                    selected["DEMONSTRACAO"] = statement
                    selected["DOCUMENTO_CVM"] = doc.upper()
                    pieces.append(selected)

    if not pieces:
        return pd.DataFrame(columns=READ_COLUMNS + ["DEMONSTRACAO"])
    return pd.concat(pieces, ignore_index=True)


def latest_filing_rows(frame: pd.DataFrame, company: Company) -> pd.DataFrame:
    if frame.empty:
        return frame

    frame = frame.copy()
    frame["DT_REFER"] = pd.to_datetime(frame["DT_REFER"], errors="coerce")
    frame["DT_FIM_EXERC"] = pd.to_datetime(frame["DT_FIM_EXERC"], errors="coerce")
    frame["VERSAO"] = pd.to_numeric(frame["VERSAO"], errors="coerce").fillna(0).astype(int)
    frame["VL_CONTA_CVM"] = pd.to_numeric(frame["VL_CONTA"], errors="coerce")
    escalas = frame["ESCALA_MOEDA"].map(normalize_text)
    escalas_desconhecidas = sorted(set(escalas.dropna()) - set(FATORES_ESCALA))
    if escalas_desconhecidas:
        raise RuntimeError(
            f"{company.ticker}: escala(s) monetaria(s) nao reconhecida(s): "
            f"{escalas_desconhecidas}"
        )
    frame["FATOR_ESCALA"] = escalas.map(FATORES_ESCALA)
    frame["VL_CONTA"] = frame["VL_CONTA_CVM"] * frame["FATOR_ESCALA"]
    frame["ORDEM_NORMALIZADA"] = frame["ORDEM_EXERC"].map(normalize_text)
    frame = frame[frame["ORDEM_NORMALIZADA"].eq("ULTIMO")]

    # Valida o mapeamento quando os dois identificadores estão disponíveis.
    unexpected = frame[
        frame["CD_CVM"].eq(company.cd_cvm) & ~frame["CNPJ_CIA"].eq(company.cnpj)
    ]
    if not unexpected.empty:
        cnpjs = ", ".join(sorted(unexpected["CNPJ_CIA"].dropna().unique()))
        raise ValueError(f"Código CVM {company.cd_cvm} associado a CNPJ inesperado: {cnpjs}")

    newest = frame.groupby("DT_REFER", dropna=False)["VERSAO"].transform("max")
    frame = frame[frame["VERSAO"].eq(newest)]
    frame = frame.dropna(subset=["DT_REFER", "CD_CONTA"])

    # Uma conta pode ser repetida por erro/reapresentação; mantém a última linha publicada.
    frame = frame.drop_duplicates(
        subset=["DT_REFER", "DEMONSTRACAO", "CD_CONTA"], keep="last"
    )
    return frame


def collect_data(zip_paths: list[tuple[str, int, Path]]) -> dict[str, pd.DataFrame]:
    results: dict[str, pd.DataFrame] = {}
    for company in COMPANIES:
        company_parts: list[pd.DataFrame] = []
        for doc, year, archive in sorted(zip_paths, key=lambda item: (item[1], item[0])):
            for statement in ("BPA", "BPP"):
                part = read_company_statement(archive, year, company, statement, doc)
                if not part.empty:
                    company_parts.append(part)
        combined = (
            pd.concat(company_parts, ignore_index=True)
            if company_parts
            else pd.DataFrame(columns=READ_COLUMNS + ["DEMONSTRACAO"])
        )
        results[company.ticker] = latest_filing_rows(combined, company)
        logging.info("%s: %d linhas selecionadas.", company.ticker, len(results[company.ticker]))
    return results


def account_sort_key(code: str) -> tuple:
    parts = re.split(r"[.\-]", str(code))
    return tuple((0, int(p)) if p.isdigit() else (1, p) for p in parts)


def section_fill(code: str) -> str | None:
    depth = str(code).count(".")
    if depth == 0:
        return NAVY
    if depth == 1:
        return LIGHT_BLUE
    return None


def add_cover_sheet(wb: Workbook, years: list[int], output_name: str) -> None:
    ws = wb.active
    ws.title = "Capa"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Balanços patrimoniais — ITR/CVM"
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:F2")
    ws["A1"].alignment = Alignment(vertical="center")

    rows = [
        ("Ficheiro", output_name),
        ("Período", f"{min(years)}–{max(years)}"),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Fonte", SOURCE_PAGE),
        ("Critério", "Maior versão por companhia e data; apenas ORDEM_EXERC = ÚLTIMO"),
        ("Escopo", "Consolidado, exceto RDOR3 (individual)"),
    ]
    for row, (label, value) in enumerate(rows, start=4):
        ws.cell(row, 1, label).font = Font(bold=True, color=NAVY)
        ws.cell(row, 2, value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws["A12"] = "Legenda"
    ws["A12"].font = Font(bold=True, color=WHITE)
    ws["A12"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A12:F12")
    legends = [
        ("Azul-escuro", "Conta total da demonstração"),
        ("Azul-claro", "Grupo principal de contas"),
        ("Verde", "Cobertura encontrada"),
        ("Vermelho-claro", "Sem ITR encontrado no intervalo"),
    ]
    for row, (label, description) in enumerate(legends, start=13):
        ws.cell(row, 1, label)
        ws.cell(row, 2, description)
    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 18


def master_account_key(row: object) -> tuple[str, str, str, str]:
    """Cria uma chave comparável sem misturar contas não fixas distintas."""
    statement = str(getattr(row, "DEMONSTRACAO"))
    code = str(getattr(row, "CD_CONTA"))
    fixed = normalize_text(getattr(row, "ST_CONTA_FIXA", "")) == "S"
    description_key = "" if fixed else normalize_text(getattr(row, "DS_CONTA", ""))
    return statement, code, "FIXA" if fixed else "NAO_FIXA", description_key


def build_master_accounts(data: dict[str, pd.DataFrame]) -> list[dict]:
    """Monta a união ordenada de contas usada, sem exceção, em todas as abas."""
    candidates: dict[tuple[str, str, str, str], list[tuple[pd.Timestamp, int, str]]] = {}
    for frame in data.values():
        if frame.empty:
            continue
        for row in frame.itertuples():
            key = master_account_key(row)
            candidates.setdefault(key, []).append(
                (pd.Timestamp(row.DT_REFER), int(row.VERSAO), str(row.DS_CONTA))
            )

    accounts = []
    for key, descriptions in candidates.items():
        statement, code, account_type, _ = key
        # A descrição mais recente prevalece nas contas fixas; nas não fixas, a
        # chave já contém a descrição normalizada e esta escolha preserva acentos.
        description = max(descriptions, key=lambda item: (item[0], item[1]))[2]
        accounts.append(
            {
                "key": key,
                "statement": statement,
                "code": code,
                "description": description,
                "account_type": account_type,
            }
        )

    return sorted(
        accounts,
        key=lambda item: (
            0 if item["statement"] == "BPA" else 1,
            account_sort_key(item["code"]),
            item["description"],
        ),
    )


def add_company_sheet(
    wb: Workbook,
    company: Company,
    frame: pd.DataFrame,
    master_accounts: list[dict],
) -> dict:
    ws = wb.create_sheet(company.ticker)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C7"
    ws["A1"] = f"{company.ticker} — Balanço Patrimonial ({company.scope_label})"
    ws["A1"].font = Font(name="Aptos Display", size=15, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:H2")
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A3"] = "Código CVM"
    ws["B3"] = company.cd_cvm
    ws["C3"] = "CNPJ"
    ws["D3"] = company.cnpj
    ws["E3"] = "Escopo"
    ws["F3"] = company.scope_label

    names = sorted(frame["DENOM_CIA"].dropna().astype(str).unique()) if not frame.empty else []
    currencies = sorted(frame["MOEDA"].dropna().astype(str).unique()) if not frame.empty else []
    scales = sorted(frame["ESCALA_MOEDA"].dropna().astype(str).unique()) if not frame.empty else []
    unit = f"{'/'.join(currencies)} — {'/'.join(scales)}"
    ws["A4"] = "Denominação CVM"
    ws["B4"] = " | ".join(names)
    ws.merge_cells("B4:F4")
    ws["G3"] = "Unidade"
    ws["H3"] = unit

    periods = sorted(frame["DT_REFER"].dropna().unique())
    period_index = {pd.Timestamp(period): index for index, period in enumerate(periods)}
    values = {
        (master_account_key(row), pd.Timestamp(row.DT_REFER)): row.VL_CONTA
        for row in frame.itertuples()
    }

    header_row = 6
    headers = ["Código CVM", "Descrição da conta"] + [pd.Timestamp(p).to_pydatetime() for p in periods]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, value)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="right" if col > 2 else "left")
        cell.border = Border(bottom=THIN_GREY)
        if col > 2:
            cell.number_format = "dd/mm/yyyy"

    row_number = header_row + 1
    previous_statement = None
    total_cells = []
    for account in master_accounts:
        statement = account["statement"]
        code = account["code"]
        account_key = account["key"]
        if statement != previous_statement:
            label = "ATIVO" if statement == "BPA" else "PASSIVO E PATRIMÔNIO LÍQUIDO"
            ws.cell(row_number, 1, label)
            ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=len(headers))
            cell = ws.cell(row_number, 1)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(bold=True, color=WHITE)
            row_number += 1
            previous_statement = statement

        ws.cell(row_number, 1, code)
        ws.cell(row_number, 2, account["description"])
        depth = str(code).count(".")
        ws.cell(row_number, 2).alignment = Alignment(indent=min(depth, 6))
        for period, index in period_index.items():
            cell = ws.cell(row_number, 3 + index, values.get((account_key, period)))
            cell.number_format = '#,##0;[Red](#,##0);-'
            cell.alignment = Alignment(horizontal="right")

        if (statement, str(code)) in (("BPA", "1"), ("BPP", "2")):
            total_type = "Ativo" if statement == "BPA" else "Passivo + PL"
            for period, index in period_index.items():
                total_cells.append(
                    {
                        "period": period,
                        "type": total_type,
                        "cell": f"{get_column_letter(3 + index)}{row_number}",
                    }
                )

        fill_color = section_fill(code)
        if fill_color:
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row_number, col)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                if fill_color == NAVY:
                    cell.font = Font(bold=True, color=WHITE)
                else:
                    cell.font = Font(bold=True, color=BLACK)
                    cell.border = Border(top=THIN_GREY)
        row_number += 1

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row_number - 1}"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 48
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.print_title_rows = f"1:{header_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_layout_view = False

    return {
        "ticker": company.ticker,
        "status": "OK" if not frame.empty else "Sem dados",
        "dates": len(periods),
        "rows": len(master_accounts),
        "unit": unit if not frame.empty else "—",
        "total_cells": total_cells,
    }


def add_coverage_sheet(wb: Workbook, coverage: list[dict], years: list[int]) -> None:
    ws = wb.create_sheet("Cobertura", 1)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Cobertura e controlos"
    ws["A1"].font = Font(name="Aptos Display", size=15, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:F2")
    headers = ["Ticker", "Escopo", "Status", "Datas", "Contas", "Unidade"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(4, col, value)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE)
    company_by_ticker = {c.ticker: c for c in COMPANIES}
    for row, item in enumerate(coverage, 5):
        company = company_by_ticker[item["ticker"]]
        values = [item["ticker"], company.scope_label, item["status"], item["dates"], item["rows"], item["unit"]]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        ws.cell(row, 3).fill = PatternFill("solid", fgColor=GREEN if item["status"] == "OK" else RED)
    last_row = 4 + len(coverage)
    ws.auto_filter.ref = f"A4:F{last_row}"
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 24
    ws["A14"] = "Intervalo solicitado"
    ws["B14"] = f"{min(years)}–{max(years)}"
    ws["A15"] = "Nota"
    ws["B15"] = "O ano corrente pode conter somente os trimestres já entregues à CVM."
    ws.merge_cells("B15:F15")


def add_sources_sheet(wb: Workbook, zip_paths: dict[int, Path]) -> None:
    ws = wb.create_sheet("Fontes")
    ws.sheet_view.showGridLines = False
    ws.append(["Ano", "URL do ZIP", "Ficheiro local", "Fonte institucional"])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE)
    for year, path in sorted(zip_paths.items()):
        ws.append([year, BASE_URL.format(year=year), str(path.resolve()), SOURCE_PAGE])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(zip_paths) + 1}"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 74
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 62


def add_checks_sheet(wb: Workbook, coverage: list[dict]) -> None:
    ws = wb.create_sheet("Checks", 2)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Verificação: Ativo = Passivo + Patrimônio Líquido"
    ws["A1"].font = Font(name="Aptos Display", size=15, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:G2")
    headers = ["Ticker", "Data", "Ativo", "Passivo + PL", "Diferença", "Tolerância", "Status"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(4, col, value)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE)

    row = 5
    for item in coverage:
        cells_by_period: dict[pd.Timestamp, dict[str, str]] = {}
        for total in item["total_cells"]:
            cells_by_period.setdefault(pd.Timestamp(total["period"]), {})[total["type"]] = total["cell"]
        for period, cells in sorted(cells_by_period.items()):
            ws.cell(row, 1, item["ticker"])
            ws.cell(row, 2, period.to_pydatetime())
            ws.cell(row, 2).number_format = "dd/mm/yyyy"
            if "Ativo" in cells:
                ws.cell(row, 3, f"='{item['ticker']}'!{cells['Ativo']}")
            if "Passivo + PL" in cells:
                ws.cell(row, 4, f"='{item['ticker']}'!{cells['Passivo + PL']}")
            ws.cell(row, 5, f"=C{row}-D{row}")
            ws.cell(row, 6, 1)
            ws.cell(row, 7, f'=IF(AND(ISNUMBER(C{row}),ISNUMBER(D{row}),ABS(E{row})<=F{row}),"OK","VERIFICAR")')
            for col in range(3, 7):
                ws.cell(row, col).number_format = '#,##0;[Red](#,##0);-'
            row += 1

    ws.auto_filter.ref = f"A4:G{row - 1}"
    ws.freeze_panes = "A5"
    widths = [12, 14, 18, 18, 16, 14, 14]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws["A3"] = "Tolerância de 1 unidade da escala publicada para eventuais arredondamentos."
    ws.merge_cells("A3:G3")
    ws["A3"].font = Font(italic=True, color="666666")


def build_workbook(
    data: dict[str, pd.DataFrame], zip_paths: dict[int, Path], output: Path, years: list[int]
) -> list[dict]:
    wb = Workbook()
    add_cover_sheet(wb, years, output.name)
    master_accounts = build_master_accounts(data)
    if not master_accounts:
        raise RuntimeError("Nenhuma conta foi encontrada para montar a estrutura-mestre.")
    coverage = []
    for company in COMPANIES:
        coverage.append(
            add_company_sheet(wb, company, data[company.ticker], master_accounts)
        )
    add_coverage_sheet(wb, coverage, years)
    add_checks_sheet(wb, coverage)
    add_sources_sheet(wb, zip_paths)
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(".tmp.xlsx")
    wb.save(temporary)
    os.replace(temporary, output)
    return coverage


def json_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if hasattr(value, "item"):
        return value.item()
    return value


def build_company_json(
    company: Company,
    frame: pd.DataFrame,
    master_accounts: list[dict],
) -> dict:
    names = sorted(frame["DENOM_CIA"].dropna().astype(str).unique()) if not frame.empty else []
    currencies = sorted(frame["MOEDA"].dropna().astype(str).unique()) if not frame.empty else []
    scales = sorted(frame["ESCALA_MOEDA"].dropna().astype(str).unique()) if not frame.empty else []
    unit = "Reais integrais" if not frame.empty else None
    periods = [pd.Timestamp(period) for period in sorted(frame["DT_REFER"].dropna().unique())]
    values = {
        (master_account_key(row), pd.Timestamp(row.DT_REFER)): row.VL_CONTA
        for row in frame.itertuples()
    }

    rows = []
    for account in master_accounts:
        account_key = account["key"]
        rows.append(
            {
                "statement": account["statement"],
                "code": account["code"],
                "description": account["description"],
                "account_type": account["account_type"],
                "depth": str(account["code"]).count("."),
                "values": {
                    period.date().isoformat(): json_value(values.get((account_key, period)))
                    for period in periods
                },
            }
        )

    totals: dict[str, dict[str, object]] = {}
    for row in frame.itertuples():
        statement = str(row.DEMONSTRACAO)
        code = str(row.CD_CONTA)
        if (statement, code) not in (("BPA", "1"), ("BPP", "2")):
            continue
        period = pd.Timestamp(row.DT_REFER).date().isoformat()
        totals.setdefault(period, {})["ativo" if statement == "BPA" else "passivo_pl"] = json_value(row.VL_CONTA)

    checks = []
    for period, total in sorted(totals.items()):
        ativo = total.get("ativo")
        passivo_pl = total.get("passivo_pl")
        difference = ativo - passivo_pl if ativo is not None and passivo_pl is not None else None
        checks.append(
            {
                "period": period,
                "ativo": ativo,
                "passivo_pl": passivo_pl,
                "difference": difference,
                "status": "OK" if difference is not None and abs(difference) <= 1 else "VERIFICAR",
            }
        )

    return {
        "ticker": company.ticker,
        "cd_cvm": company.cd_cvm,
        "cnpj": company.cnpj,
        "scope": company.scope_label,
        "expected_name": company.expected_name,
        "denom_cvm": names,
        "moeda": currencies,
        "escala_moeda_original": scales,
        "unit": unit,
        "status": "OK" if not frame.empty else "Sem dados",
        "periods": [period.date().isoformat() for period in periods],
        "rows": rows,
        "checks": checks,
    }


def export_json(
    data: dict[str, pd.DataFrame], zip_paths: list[tuple[str, int, Path]], output: Path, years: list[int]
) -> list[dict]:
    master_accounts = build_master_accounts(data)
    if not master_accounts:
        raise RuntimeError("Nenhuma conta foi encontrada para montar a estrutura-mestre.")

    companies = {
        company.ticker: build_company_json(company, data[company.ticker], master_accounts)
        for company in COMPANIES
    }
    coverage = [
        {
            "ticker": ticker,
            "scope": payload["scope"],
            "status": payload["status"],
            "dates": len(payload["periods"]),
            "rows": len(payload["rows"]),
            "unit": payload["unit"],
        }
        for ticker, payload in companies.items()
    ]
    payload = {
        "kind": "balanco_patrimonial_itr_cvm",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": SOURCE_PAGE,
        "dfp_source": DFP_SOURCE_PAGE,
        "years": years,
        "criteria": "ITR trimestral e DFP anual quando disponivel; maior versao por companhia e data; apenas ORDEM_EXERC = ULTIMO",
        "scope_note": "Consolidado, exceto RDOR3 (individual)",
        "coverage": coverage,
        "sources": [
            {
                "document": doc.upper(),
                "year": year,
                "zip_url": (DFP_BASE_URL if doc == "dfp" else BASE_URL).format(year=year),
                "local_file": str(path.resolve()),
                "source_page": DFP_SOURCE_PAGE if doc == "dfp" else SOURCE_PAGE,
            }
            for doc, year, path in sorted(zip_paths, key=lambda item: (item[1], item[0]))
        ],
        "companies": companies,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(".tmp.json")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, output)
    return coverage


def verify_json(output: Path, coverage: list[dict]) -> None:
    payload = json.loads(output.read_text(encoding="utf-8"))
    expected = {c.ticker for c in COMPANIES}
    found = set(payload.get("companies", {}))
    missing = expected.difference(found)
    if missing:
        raise RuntimeError(f"Empresas ausentes no JSON: {sorted(missing)}")
    if len(payload.get("coverage", [])) != len(COMPANIES):
        raise RuntimeError("A cobertura do JSON esta incompleta.")
    company_rows = {item["ticker"]: len(payload["companies"][item["ticker"]]["rows"]) for item in coverage}
    if len(set(company_rows.values())) != 1:
        raise RuntimeError(f"As empresas nao tem a mesma estrutura de linhas: {company_rows}")
    if output.stat().st_size < 1_000:
        raise RuntimeError("O JSON gerado parece incompleto.")


def verify_workbook(output: Path, coverage: list[dict]) -> None:
    wb = load_workbook(output, data_only=False, read_only=True)
    expected = {"Capa", "Cobertura", "Checks", "Fontes", *(c.ticker for c in COMPANIES)}
    missing = expected.difference(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Abas ausentes no Excel: {sorted(missing)}")
    if wb["Cobertura"].max_row < len(COMPANIES) + 4:
        raise RuntimeError("A aba Cobertura está incompleta.")
    if wb["Checks"].max_row <= 4:
        raise RuntimeError("A aba Checks não contém verificações.")
    for item in coverage:
        ws = wb[item["ticker"]]
        if item["status"] == "OK" and ws.max_column < 3:
            raise RuntimeError(f"Aba {item['ticker']} sem colunas de períodos.")
    company_rows = {item["ticker"]: wb[item["ticker"]].max_row for item in coverage}
    if len(set(company_rows.values())) != 1:
        raise RuntimeError(f"As abas não têm a mesma estrutura de linhas: {company_rows}")
    wb.close()
    if output.stat().st_size < 10_000:
        raise RuntimeError("O Excel gerado parece incompleto.")


def main() -> int:
    global COMPANIES
    args = parse_args()
    COMPANIES = financial_companies(args.sector)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )
    downloads = args.output_dir / "downloads"
    try:
        zip_paths = [
            ("itr", year, download_zip(year, downloads / "itr", args.force_download, args.offline, "itr"))
            for year in args.years
        ]
        if not args.no_dfp:
            for year in args.years:
                try:
                    zip_paths.append(
                        ("dfp", year, download_zip(year, downloads / "dfp", args.force_download, args.offline, "dfp"))
                    )
                except Exception as exc:
                    logging.warning("DFP %s nao incorporado: %s", year, exc)
        data = collect_data(zip_paths)
        output = args.output_dir / f"balancos_itr_cvm_{min(args.years)}_{max(args.years)}.json"
        coverage = export_json(data, zip_paths, output, args.years)
        verify_json(output, coverage)
    except Exception as exc:
        logging.error("Execução interrompida: %s", exc)
        return 1

    logging.info("Concluído: %s", output.resolve())
    for item in coverage:
        logging.info(
            "%s | %s | %d datas | %d contas",
            item["ticker"], item["status"], item["dates"], item["rows"],
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
