"""Baixa planilhas de fundamentos de companhias de saúde e exporta KPIs em JSON.

Dependências:
    pip install openpyxl playwright
    playwright install chromium

Uso online (baixa a planilha mais recente e cria um JSON por ticker):
    python ri_fundamentos.py

Uso local (sem acessar os sites; útil para teste/contingência):
    python ri_fundamentos.py --file RDOR3=/caminho/RDOR3.xlsx \
        --file FLRY3=/caminho/FLRY3.xlsm

Por padrão os JSONs são gravados no diretório atual. Use --output-dir para mudar.

Princípio importante: o programa busca apenas Ticket Médio, N. Atendimentos,
N. Unidades, N. Pacientes, Receita Bruta e Glosa/PCLD. Ele não converte
métricas diferentes em equivalentes, salvo proxies explicitamente definidos no
dicionário operacional. Quando "Pacientes-Dia" alimenta N. Atendimentos ou
N. Pacientes em MATD3/RDOR3, a saída preserva natureza de proxy e confiança
média. Quando o indicador solicitado não é divulgado na planilha, a saída
contém uma lista vazia para aquele indicador.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import math
import re
import sys
import tempfile
import unicodedata
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook

from operational_dictionary import (
    CONFIDENCE_HIGH,
    CONFIDENCE_MEDIUM,
    all_metric_names,
    metric_aliases,
    metric_definition,
)

try:
    from app_parser_operacional import (
        PASTA_ENTRADA_PADRAO as PASTA_PDFS_OPERACIONAIS,
        PASTA_SAIDA_PADRAO as PASTA_MARKDOWNS_OPERACIONAIS,
        converter_pdf_para_markdown,
        listar_pdfs_entrada,
        normalizar_nome_arquivo,
    )
except Exception:  # pragma: no cover - fallback defensivo para ambientes sem parser.
    PASTA_PDFS_OPERACIONAIS = Path(__file__).resolve().parent / "Releases e relatórios" / "entrada"
    PASTA_MARKDOWNS_OPERACIONAIS = Path(__file__).resolve().parent / "Releases e relatórios" / "saída"
    converter_pdf_para_markdown = None
    listar_pdfs_entrada = None

    def normalizar_nome_arquivo(nome: str) -> str:
        text = unicodedata.normalize("NFKD", nome)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = re.sub(r"[^A-Za-z0-9._-]+", "_", text)
        return text.strip("._-") or "documento"


METRIC_NAMES = all_metric_names()


@dataclass(frozen=True)
class Company:
    ticker: str
    name: str
    results_url: str
    fallback_pages: tuple[str, ...] = ()
    preferred_sheets: tuple[str, ...] = ()
    sheet_aliases: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class SnapshotCell:
    value: Any


@dataclass
class SheetSnapshot:
    title: str
    rows: list[tuple[Any, ...]]

    @property
    def max_row(self) -> int:
        return len(self.rows)

    @property
    def max_column(self) -> int:
        return max((len(row) for row in self.rows), default=0)

    def cell(self, row: int, col: int) -> SnapshotCell:
        if row < 1 or col < 1 or row > len(self.rows):
            return SnapshotCell(None)
        values = self.rows[row - 1]
        if col > len(values):
            return SnapshotCell(None)
        return SnapshotCell(values[col - 1])


@dataclass
class WorkbookSnapshot:
    worksheets: list[SheetSnapshot]


COMPANIES: dict[str, Company] = {
    "RDOR3": Company(
        "RDOR3",
        "Rede D'Or São Luiz",
        "https://ri.rededorsaoluiz.com.br/informacoes-financeiras/central-de-resultados/",
        ("https://ri.rededorsaoluiz.com.br/informacoes-financeiras/planilha-de-fundamentos/",),
        (r"^Portugu[eê]s$",),
    ),
    "MATD3": Company(
        "MATD3",
        "Rede Mater Dei",
        "https://ri.materdei.com.br/informacoes-aos-acionistas/central-de-resultados/",
        (),
        (r"^Portugu[eê]s$",),
    ),
    "FLRY3": Company(
        "FLRY3",
        "Grupo Fleury",
        "https://ri.fleury.com.br/informacoes-financeiras-e-apresentacoes/central-de-resultados/",
        ("https://ri.fleury.com.br/informacoes-financeiras-e-apresentacoes/planilha-de-dados-historicos/",),
        (r"^Combinada Outras Informa",),
    ),
    "AALR3": Company(
        "AALR3",
        "Alliança Saúde",
        "https://ri.allianca.com/informacoes-financeiras/central-de-resultados/",
        ("https://ri.allianca.com/informacoes-financeiras/dados-historicos/",),
    ),
    "DASA3": Company(
        "DASA3",
        "Dasa",
        "https://www.dasa3.com.br/informacoes-financeiras/resultado-trimestral/",
        ("https://www.dasa3.com.br/servicos-aos-investidores/central-downloads/",),
        (r"^2$", r"^3$"),
        {
            "1": "Consolidado",
            "2": "Diagnósticos Nacional",
            "3": "Hospitais/Onco NE",
            "4": "Américas",
        },
    ),
    "ONCO3": Company(
        "ONCO3",
        "Grupo Oncoclínicas",
        "https://ri.grupooncoclinicas.com/informacoes-financeiras/central-de-resultados/",
        ("https://ri.grupooncoclinicas.com/servicos-aos-investidores/central-de-downloads/",),
        (r"^DRE Trimestral$",),
    ),
    "HAPV3": Company(
        "HAPV3",
        "Hapvida",
        "https://ri.hapvida.com.br/informacoes-financeiras/central-de-resultados/",
        ("https://ri.hapvida.com.br/servicos-aos-investidores/central-de-downloads/",),
    ),
}


GENERIC_PATTERNS: dict[str, tuple[str, ...]] = {
    "Ticket Médio": (r"^ticket medio(?:\b|\s|\()", r"^average ticket(?:\b|\s|\()"),
    "N. Atendimentos": (
        r"^(?:numero de )?atendimentos(?:\b|\s|\()",
        r"^volume de atendimentos",
        r"^consultas(?:\b|\s|\()",
    ),
    "N. Unidades": (
        r"^(?:numero de )?unidades(?:\b|\s|\()",
        r"^unidades de atendimento",
        r"^unidades operacionais",
        r"^unidades proprias",
    ),
    "N. Pacientes": (
        r"^(?:numero de )?pacientes(?:\b|\s|\()",
        r"^pacientes oncol",
    ),
    "Receita Bruta": (r"^receita bruta$", r"^gross revenue$"),
    "Glosa/PCLD": (
        r"^glosas?$",
        r"^pcld$",
        r"^impostos, deducoes e glosas",
        r"^provisao .*creditos de liquidacao duvidosa",
        r"^constituicao .*provisao para creditos de liquidacao duvidosa",
        r"^constituicao .*provisao para glosas",
        r"^perdas estimadas para glosa e creditos de liquidacao duvidosa",
        r"^provisao .*perda de credito esperada e glosas",
        r"^allowance for doubtful accounts$",
    ),
}

GENERIC_PATTERNS.update(
    {
        metric: tuple(rf"^{re.escape(alias)}(?:\b|\s|\(|$)" for alias in metric_aliases("", metric))
        for metric in METRIC_NAMES
        if metric not in GENERIC_PATTERNS
    }
)


# Regras que melhoram a precisão quando a nomenclatura da companhia é específica.
COMPANY_PATTERNS: dict[str, dict[str, tuple[str, ...]]] = {
    "RDOR3": {
        "N. Unidades": (r"^numero de hospitais proprios em operacao", r"^hospitais proprios"),
        "N. Atendimentos": (r"^pacientes[- ]dia$", r"^oncologia - infusoes", r"^infusoes"),
        "N. Pacientes": (r"^pacientes[- ]dia$",),
        "Ticket Médio": (r"^ticket medio$",),
        "Glosa/PCLD": (r"^glosas?$",),
    },
    # No Fleury, "Número de Unidades" é um cabeçalho com zeros. O total de
    # unidades diagnósticas fica na linha imediatamente abaixo, "Medicina Diagnóstica".
    "FLRY3": {
        "N. Unidades": (r"^medicina diagnostica$",),
        "N. Atendimentos": (r"^atendimentos(?:\b|\s|\()",),
        "N. Pacientes": (r"^atendimentos(?:\b|\s|\()",),
        "Glosa/PCLD": (r"^glosas?$", r"^glosas e abatimentos$"),
    },
    "MATD3": {
        "Ticket Médio": (r"^ticket medio(?:\b|\s|\()",),
        "N. Atendimentos": (r"^pacientes[- ]dia$",),
        "N. Pacientes": (r"^pacientes oncol", r"^pacientes"),
        "Receita Bruta": (r"^receita bruta$",),
        "Glosa/PCLD": (r"^constituicao .*provisao para glosas$", r"^glosas?$"),
    },
    "HAPV3": {
        "Ticket Médio": (r"^ticket medio .*saude", r"^ticket medio \\(saude\\)"),
        "N. Unidades": (r"^unidades da rede propria", r"^unidades(?:\b|\s|\()", r"^rede propria"),
        "Glosa/PCLD": (r"^provisao.*glosa esperada",),
    },
    "DASA3": {
        "Ticket Médio": (r"^ticket medio \(r\$\)", r"^ticket medio$",),
        "N. Atendimentos": (r"^exames - total", r"^exames total"),
        "N. Unidades": (r"^unidades de atendimento$",),
    },
    "ONCO3": {
        "N. Atendimentos": (r"^total de procedimentos$", r"^procedimentos$", r"^infusoes$"),
        "N. Pacientes": (r"^total de procedimentos$", r"^procedimentos$"),
        "N. Unidades": (r"^numero de unidades$",),
        "Ticket Médio": (r"^ticket medio$",),
        "Glosa/PCLD": (r"^pcld$",),
    },
}


# Para métricas financeiras, a aba mais adequada nem sempre é a mesma usada
# para KPIs operacionais. Ex.: no Fleury, glosas ficam na DRE/DFC combinada.
FINANCIAL_SHEET_PATTERNS: dict[str, dict[str, tuple[str, ...]]] = {
    "RDOR3": {
        "Receita Bruta": (r"^Portugu[eê]s$",),
        "Glosa/PCLD": (r"^Portugu[eê]s$",),
    },
    "MATD3": {
        "Receita Bruta": (r"^Portugu[eê]s$",),
        "Glosa/PCLD": (r"^Portugu[eê]s$",),
    },
    "FLRY3": {
        "Receita Bruta": (r"^Combinada DRE$",),
        "Glosa/PCLD": (r"^Combinada (?:DRE|DFC)$",),
    },
    "DASA3": {
        "Receita Bruta": (r"^1$",),
        "Glosa/PCLD": (r"^1$",),
    },
    "ONCO3": {
        "Receita Bruta": (r"^DRE Trimestral$",),
        "Glosa/PCLD": (r"^DRE Trimestral$",),
    },
}


FINANCIAL_UNITS: dict[str, str] = {
    "AALR3": "R$ milhões",
    "RDOR3": "R$ milhões",
    "MATD3": "R$ milhares",
    "FLRY3": "R$ milhares",
    "DASA3": "R$ milhões",
    "ONCO3": "R$ milhões",
}

COMPANY_MARKDOWN_ALIASES: dict[str, tuple[str, ...]] = {
    "RDOR3": ("rdor3", "rede dor", "rede d'or", "rededor"),
    "MATD3": ("matd3", "mater dei"),
    "FLRY3": ("flry3", "fleury"),
    "AALR3": ("aalr3", "allianca", "alliança"),
    "DASA3": ("dasa3", "dasa"),
    "ONCO3": ("onco3", "oncoclinicas", "oncoclínicas"),
    "HAPV3": ("hapv3", "hapvida"),
}


def metric_unit(company: Company, metric: str, source_label: str) -> str | None:
    if metric in {"Receita Bruta", "Glosa/PCLD"}:
        return FINANCIAL_UNITS.get(company.ticker)
    if metric == "N. Unidades":
        return "unidades"
    if metric == "N. Pacientes":
        return "pacientes"
    if metric == "N. Atendimentos":
        return "mil atendimentos" if company.ticker == "FLRY3" else "atendimentos"
    if metric == "Ticket Médio":
        if company.ticker == "MATD3":
            return "R$ milhões por leito utilizado"
        return "R$"
    return None


def normalise_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def confidence_label(score: int) -> str:
    if score >= CONFIDENCE_HIGH:
        return "high"
    if score >= CONFIDENCE_MEDIUM:
        return "medium"
    return "low"


def infer_observation_nature(
    company: Company,
    metric: str,
    label: str,
    context: str,
    *,
    calculated: bool = False,
) -> tuple[str, str | None]:
    if calculated:
        return "calculated", "formula"
    definition = metric_definition(company.ticker, metric)
    haystack = normalise_text(f"{label} {context}")
    for proxy in definition.get("allowed_proxies", ()):
        if normalise_text(proxy) in haystack:
            return "proxy", proxy
    return "reported", None


def forbidden_operational_context(company: Company, metric: str, label: str, context: str) -> str | None:
    haystack = normalise_text(f"{label} {context}")
    definition = metric_definition(company.ticker, metric)
    allowed_proxies = tuple(normalise_text(proxy) for proxy in definition.get("allowed_proxies", ()))
    for term in definition.get("forbidden_contexts", ()):
        normalized = normalise_text(term)
        if normalized in allowed_proxies:
            continue
        if normalized and normalized in haystack:
            return term
    if metric == "Ticket Médio":
        if re.search(r"\b(?:var\.?|variacao|yoy|qoq|a/a|t/t)\b", haystack):
            return "variation_context"
    if metric == "N. Pacientes" and any(term in haystack for term in ("pacientes-dia", "paciente-dia")):
        if any(proxy in haystack for proxy in allowed_proxies):
            return None
        return "pacientes-dia_is_not_unique_patients"
    if metric == "Receita Bruta" and any(term in haystack for term in ("receita liquida", "net revenue")):
        return "net_revenue_is_not_gross_revenue"
    return None


def classify_operational_observation(
    company: Company,
    metric: str,
    *,
    label: str,
    value: Any = None,
    unit: str | None = None,
    period: str | None = None,
    scope: str | None = None,
    context: str = "",
    document: str | None = None,
    page: int | None = None,
    extraction_method: str = "",
    calculated: bool = False,
    formula: str | None = None,
    inputs: dict[str, Any] | None = None,
    source_type: str | None = None,
    sheet: str | None = None,
) -> dict[str, Any]:
    rejection = forbidden_operational_context(company, metric, label, context)
    nature, proxy = infer_observation_nature(company, metric, label, context, calculated=calculated)
    mapping_type = "exact_label" if nature == "reported" else nature
    source_score = 90 if extraction_method.startswith("spreadsheet") else 86 if extraction_method == "release_table" else 76
    score = source_score
    if calculated:
        score = 82
    if nature == "proxy":
        score = 82 if company.ticker == "DASA3" and metric == "N. Atendimentos" else 74
    if (
        metric == "Receita Bruta"
        and is_number(value)
        and float(value) < 10
        and (unit is None or "r$" in normalise_text(unit))
    ):
        rejection = rejection or "gross_revenue_value_incompatible_with_unit"
        score = 35
    if rejection:
        score = 35
    confidence = confidence_label(score)
    requires_review = confidence == "low" or bool(rejection)
    return {
        "metric": metric,
        "nature": nature,
        "original_label": label,
        "original_value": serialisable_number(value) if is_number(value) else value,
        "normalized_value": serialisable_number(value) if is_number(value) else value,
        "unit": unit,
        "period": period,
        "period_type": "quarter" if period and re.fullmatch(r"[1-4]T\d{2}", period) else "annual" if period else None,
        "measurement_basis": "period_value",
        "scope": scope,
        "document": document,
        "sheet": sheet,
        "page": page,
        "source_context": context[:500],
        "original_metric": proxy.title() if proxy else label,
        "original_unit": unit,
        "normalized_unit": unit,
        "mapping_type": mapping_type,
        "proxy_for": proxy,
        "formula": formula,
        "inputs": inputs or {},
        "source_type": source_type or extraction_method,
        "source_confidence": confidence_label(source_score),
        "source_confidence_score": source_score,
        "warning": operational_warning_message(company, metric, {"fonte_linha": label}) if nature == "proxy" else None,
        "confidence": confidence,
        "confidence_score": score,
        "requires_review": requires_review,
        "rejection_reason": rejection,
        "status": "low_confidence_rejected" if rejection else "validated",
        "extraction_method": extraction_method,
    }


def enrich_metric_item(
    company: Company,
    metric: str,
    item: dict[str, Any],
    *,
    context: str = "",
) -> dict[str, Any] | None:
    series = item.get("serie") or {}
    label = str(item.get("fonte_linha") or item.get("original_label") or metric)
    scope = str(item.get("escopo") or item.get("scope") or "")
    method = str(item.get("extraction_method") or "")
    calculated = bool(item.get("calculado"))
    rejection = forbidden_operational_context(company, metric, label, f"{scope} {context}")
    observations = [
        classify_operational_observation(
            company,
            metric,
            label=label,
            value=value,
            unit=item.get("unidade"),
            period=period,
            scope=scope,
            context=f"{scope} {context}",
            document=item.get("fonte_documento"),
            page=item.get("page"),
            extraction_method=method,
            calculated=calculated,
            formula=item.get("formula"),
            inputs=item.get("inputs"),
            source_type=item.get("source_type"),
            sheet=item.get("sheet"),
        )
        for period, value in series.items()
    ]
    low_observation = next((obs for obs in observations if obs.get("confidence") == "low"), None)
    if rejection or low_observation:
        item["confidence"] = "low"
        item["confidence_score"] = 35
        item["requires_review"] = True
        item["rejection_reason"] = rejection or low_observation.get("rejection_reason") or "low_confidence_observation"
        item["observations"] = observations
        return item
    score = min((obs["confidence_score"] for obs in observations), default=90)
    nature = "calculated" if calculated else next((obs["nature"] for obs in observations if obs["nature"] != "reported"), "reported")
    item.setdefault("nature", nature)
    item.setdefault("mapping_type", "formula" if calculated else ("proxy" if nature == "proxy" else "exact_label"))
    item["confidence_score"] = score
    item["confidence"] = confidence_label(score)
    item["requires_review"] = score < CONFIDENCE_MEDIUM
    item["observations"] = observations
    return item


def metric_item_rank(item: dict[str, Any]) -> tuple[int, int, int, str]:
    confidence_rank = {"high": 0, "medium": 1, "low": 9}.get(str(item.get("confidence")), 5)
    nature_rank = {"reported": 0, "calculated": 1, "proxy": 2}.get(str(item.get("nature")), 5)
    method = str(item.get("extraction_method", ""))
    source_rank = 0 if method.startswith("spreadsheet") else 1 if method == "release_table" else 2
    label = normalise_text(item.get("fonte_linha"))
    return confidence_rank, nature_rank, source_rank, label


def normalise_period(value: Any) -> str | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and float(value).is_integer():
        year = int(value)
        if 2000 <= year <= 2100:
            return str(year)
    if not isinstance(value, str):
        return None
    text = value.strip().upper().replace(" ", "")
    match = re.fullmatch(r"([1-4])[TQ](\d{2}|\d{4})", text)
    if match:
        quarter, year = match.groups()
        year = year[-2:]
        return f"{quarter}T{year}"
    if re.fullmatch(r"20\d{2}", text):
        return text
    return None


def period_sort_key(period: str) -> tuple[int, int]:
    if re.fullmatch(r"20\d{2}", period):
        return int(period), 5
    match = re.fullmatch(r"([1-4])T(\d{2})", period)
    if not match:
        return 0, 0
    quarter, yy = match.groups()
    return 2000 + int(yy), int(quarter)


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def serialisable_number(value: int | float) -> int | float:
    number = float(value)
    if number.is_integer():
        return int(number)
    return number


def safe_print(message: Any = "") -> None:
    text = str(message)
    encoding = getattr(sys.stdout, "encoding", None) or "utf-8"
    print(text.encode(encoding, errors="replace").decode(encoding, errors="replace"))


def sheet_is_preferred(company: Company, sheet_name: str) -> bool:
    if not company.preferred_sheets:
        return True
    return any(re.search(pattern, sheet_name, re.IGNORECASE) for pattern in company.preferred_sheets)


def metric_sheet_patterns(company: Company, metric: str) -> tuple[str, ...]:
    financial = FINANCIAL_SHEET_PATTERNS.get(company.ticker, {}).get(metric)
    if financial is not None:
        return financial
    return company.preferred_sheets


def sheet_matches_patterns(sheet_name: str, patterns: tuple[str, ...]) -> bool:
    if not patterns:
        return True
    return any(re.search(pattern, sheet_name, re.IGNORECASE) for pattern in patterns)


def is_period_header_row(ws: Any, row: int) -> bool:
    """Evita confundir valores numéricos como 2025 com um cabeçalho de ano."""
    count = 0
    for col in range(1, ws.max_column + 1):
        if normalise_period(ws.cell(row, col).value):
            count += 1
            if count >= 2:
                return True
    return False


def nearest_period_above(ws: Any, row: int, col: int, lookback: int = 500) -> str | None:
    """Procura o cabeçalho de período mais próximo na mesma coluna."""
    for candidate_row in range(row - 1, max(0, row - lookback), -1):
        period = normalise_period(ws.cell(candidate_row, col).value)
        if period and is_period_header_row(ws, candidate_row):
            return period
    return None


def row_series(ws: Any, row: int) -> dict[str, int | float]:
    series: dict[str, int | float] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if not is_number(value):
            continue
        period = nearest_period_above(ws, row, col)
        if period:
            series[period] = serialisable_number(value)
    return dict(sorted(series.items(), key=lambda item: period_sort_key(item[0])))


def has_context_above(ws: Any, row: int, pattern: str, lookback: int = 8) -> bool:
    rx = re.compile(pattern)
    for candidate_row in range(row - 1, max(0, row - lookback), -1):
        for col in range(1, min(ws.max_column, 6) + 1):
            if rx.search(normalise_text(ws.cell(candidate_row, col).value)):
                return True
    return False


def context_above_text(ws: Any, row: int, lookback: int = 12) -> str:
    values: list[str] = []
    for candidate_row in range(row - 1, max(0, row - lookback), -1):
        for col in range(1, min(ws.max_column, 8) + 1):
            raw = ws.cell(candidate_row, col).value
            if raw is not None:
                values.append(str(raw))
    return normalise_text(" ".join(values))


def rdor_hospital_scope_for_row(ws: Any, row: int) -> str | None:
    context = context_above_text(ws, row, lookback=18)
    if any(term in context for term in ("sulamerica", "sul america", "seguros e previdencia")):
        return None
    if all(term in context for term in ("hospitais", "oncologia")) or "hospitais, oncologia e outros" in context:
        return "Hospitais, oncologia e outros"
    return None


def plausible_operational_value(label: str, value: Any) -> bool:
    if not is_number(value):
        return False
    number = float(value)
    normalized_label = normalise_text(label)
    if "pacientes-dia" in normalized_label or "paciente-dia" in normalized_label:
        if number <= 0:
            return False
        return number > 100 and int(number) not in {2025, 2026}
    return True


def metric_patterns(ticker: str, metric: str) -> tuple[str, ...]:
    configured = tuple(rf"^{re.escape(alias)}(?:\b|\s|\(|$)" for alias in metric_aliases(ticker, metric))
    patterns = COMPANY_PATTERNS.get(ticker, {}).get(metric, GENERIC_PATTERNS[metric])
    return tuple(dict.fromkeys((*patterns, *configured)))


def candidate_rows(ws: Any, patterns: Iterable[str]) -> list[tuple[int, str]]:
    regexes = [re.compile(pattern) for pattern in patterns]
    found: list[tuple[int, str]] = []
    # Rótulos operacionais ficam normalmente nas primeiras colunas. Limitar a
    # busca reduz bastante o custo em planilhas históricas muito largas.
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column, 8) + 1):
            raw = ws.cell(row, col).value
            if not isinstance(raw, str):
                continue
            label = normalise_text(raw)
            if any(rx.search(label) for rx in regexes):
                found.append((row, raw.strip()))
                break
    return found


def extract_metric(company: Company, workbook: Any, metric: str) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    preferred_patterns = metric_sheet_patterns(company, metric)
    preferred_available = any(
        sheet_matches_patterns(ws.title, preferred_patterns) for ws in workbook.worksheets
    )

    for ws in workbook.worksheets:
        if preferred_available and preferred_patterns and not sheet_matches_patterns(ws.title, preferred_patterns):
            continue

        for row, source_label in candidate_rows(ws, metric_patterns(company.ticker, metric)):
            # Proteção específica do Fleury: "Medicina Diagnóstica" aparece em
            # outros blocos. Só aceitamos a ocorrência dentro de Número de Unidades.
            if company.ticker == "FLRY3" and metric == "N. Unidades":
                if not has_context_above(ws, row, r"^numero de unidades$", lookback=5):
                    continue

            # A planilha da Rede D'Or possui DRE consolidada, eliminações e
            # DRE do bloco hospitalar/oncológico na mesma aba. Por definição
            # deste projeto, Receita Bruta vem de "Hospitais, oncologia e outros".
            if company.ticker == "RDOR3" and metric == "Receita Bruta":
                if not has_context_above(ws, row, r"^hospitais, oncologia e outros$", lookback=12):
                    continue
            if company.ticker == "RDOR3" and metric == "Glosa/PCLD":
                if not has_context_above(ws, row, r"^consolidado$", lookback=12):
                    continue
            rdor_scope = None
            if company.ticker == "RDOR3" and metric in {"N. Atendimentos", "N. Pacientes", "Ticket Médio"}:
                rdor_scope = rdor_hospital_scope_for_row(ws, row)
                if rdor_scope is None:
                    continue

            series = {
                period: value
                for period, value in row_series(ws, row).items()
                if plausible_operational_value(source_label, value)
            }
            if not series:
                continue
            scope = rdor_scope or company.sheet_aliases.get(ws.title, ws.title)
            item = {
                "escopo": scope,
                "fonte_linha": source_label,
                "sheet": ws.title,
                "unidade": metric_unit(company, metric, source_label),
                "calculado": False,
                "confidence": "high",
                "extraction_method": "spreadsheet_labeled_row",
                "requires_review": False,
                "serie": series,
            }
            enriched = enrich_metric_item(company, metric, item, context=scope)
            if enriched is not None:
                results.append(enriched)

    # Remove séries repetidas, preservando a primeira ocorrência.
    unique: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in results:
        signature = json.dumps(item["serie"], ensure_ascii=False, sort_keys=True)
        if signature not in seen:
            seen.add(signature)
            unique.append(item)
    return sorted(unique, key=metric_item_rank)


def extract_arbitrary_row(
    company: Company,
    workbook: Any,
    patterns: tuple[str, ...],
    preferred_sheet_pattern: str,
) -> dict[str, int | float] | None:
    for ws in workbook.worksheets:
        if not re.search(preferred_sheet_pattern, ws.title, re.IGNORECASE):
            continue
        for row, _label in candidate_rows(ws, patterns):
            series = row_series(ws, row)
            if series:
                return series
    return None


def derive_fleury_average_ticket(company: Company, workbook: Any) -> list[dict[str, Any]]:
    """Receita bruta (R$ mil) / atendimentos (mil) = R$ por atendimento."""
    revenue = extract_arbitrary_row(
        company,
        workbook,
        (r"^receita bruta$",),
        r"^Combinada Outras Informa",
    )
    appointments = extract_arbitrary_row(
        company,
        workbook,
        (r"^atendimentos$",),
        r"^Combinada Outras Informa",
    )
    if not revenue or not appointments:
        return []
    common = sorted(set(revenue) & set(appointments), key=period_sort_key)
    series = {
        period: revenue[period] / appointments[period]
        for period in common
        if appointments[period] != 0
    }
    if not series:
        return []
    return [
        enrich_metric_item(
            company,
            "Ticket Médio",
            {
            "escopo": "Empresa Combinada",
            "fonte_linha": "Receita Bruta / Atendimentos",
            "unidade": "R$ por atendimento",
            "calculado": True,
            "formula": "Receita Bruta (R$ milhares) / Atendimentos (milhares)",
            "serie": series,
            },
            context="Empresa Combinada",
        )
    ]


def prune_trailing_zero_periods(metrics: dict[str, list[dict[str, Any]]]) -> None:
    """Remove colunas futuras pré-formatadas que aparecem apenas como zero."""
    nonzero_periods: list[str] = []
    for items in metrics.values():
        for item in items:
            if item.get("confidence") == "low":
                continue
            for period, value in item.get("serie", {}).items():
                if is_number(value) and float(value) != 0:
                    nonzero_periods.append(period)
    if not nonzero_periods:
        return
    latest_nonzero = max(nonzero_periods, key=period_sort_key)
    latest_key = period_sort_key(latest_nonzero)
    for items in metrics.values():
        for item in items:
            if item.get("confidence") == "low":
                continue
            series = item.get("serie", {})
            for period in list(series):
                if period_sort_key(period) > latest_key:
                    del series[period]


def operational_warning_message(company: Company, metric: str, item: dict[str, Any] | None = None) -> str:
    if item is None:
        return "Nenhuma observação suficientemente confiável foi encontrada."
    label = normalise_text(item.get("fonte_linha"))
    if company.ticker == "DASA3" and metric == "N. Atendimentos" and "exames" in label:
        return "Volume total de exames utilizado como proxy de atendimentos, refletindo a natureza predominantemente diagnóstica da operação."
    if company.ticker == "FLRY3" and metric == "N. Pacientes" and "atendimentos" in label:
        return "Atendimentos utilizados como proxy; não representa pacientes únicos."
    if company.ticker == "MATD3" and metric == "N. Atendimentos" and "pacientes-dia" in label:
        return "Pacientes-dia utilizado como proxy de atendimentos."
    if company.ticker == "MATD3" and metric == "N. Pacientes" and "pacientes-dia" in label:
        return "Pacientes-dia utilizado como proxy de pacientes; não representa pacientes únicos."
    if company.ticker == "ONCO3" and metric in {"N. Atendimentos", "N. Pacientes"}:
        return "Procedimentos utilizados como proxy; não representa pacientes únicos."
    if company.ticker == "RDOR3" and metric == "N. Unidades":
        return "Hospitais próprios utilizados como proxy de unidades."
    if company.ticker == "RDOR3" and metric == "N. Atendimentos":
        return "Pacientes-dia utilizado como proxy de atendimentos."
    if company.ticker == "RDOR3" and metric == "N. Pacientes":
        return "Pacientes-dia utilizado como proxy de pacientes; não representa pacientes únicos."
    return "Indicador exibido com confiança média por depender de proxy ou contexto menos estruturado."


def build_operational_warnings(company: Company, metrics: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    for metric in METRIC_NAMES:
        items = metrics.get(metric) or []
        valid_items = [item for item in items if item.get("confidence") != "low"]
        for item in items:
            if item.get("confidence") == "low":
                warnings.append(
                    {
                        "metric": metric,
                        "status": "low_confidence_rejected",
                        "message": "Candidato rejeitado por valor incompatível com unidade/contexto.",
                        "confidence": "low",
                        "candidate_value": next(iter((item.get("serie") or {}).values()), None),
                        "fonte_linha": item.get("fonte_linha"),
                        "escopo": item.get("escopo"),
                        "motivo": item.get("rejection_reason"),
                    }
                )
        if not valid_items:
            warnings.append(
                {
                    "metric": metric,
                    "status": "not_found",
                    "message": operational_warning_message(company, metric, None),
                    "confidence": "not_found",
                }
            )
            continue
        for item in valid_items:
            if item.get("confidence") == "medium":
                warnings.append(
                    {
                        "metric": metric,
                        "status": "medium_confidence",
                        "message": operational_warning_message(company, metric, item),
                        "confidence": item.get("confidence"),
                        "nature": item.get("nature"),
                        "fonte_linha": item.get("fonte_linha"),
                        "escopo": item.get("escopo"),
                    }
                )
    return warnings


def parse_markdown_number(raw: str) -> int | float | None:
    text = raw.strip()
    if not text or "%" in text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = re.sub(r"[^\d,.\-]", "", text)
    if not re.search(r"\d", text):
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+", text):
        text = text.replace(".", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        value = float(text)
    except ValueError:
        return None
    if negative:
        value = -value
    return serialisable_number(value)


def markdown_periods(text: str) -> list[str]:
    periods: list[str] = []
    for quarter, year in re.findall(r"\b([1-4])\s*[TQ]\s*(\d{2}|\d{4})\b", text, re.IGNORECASE):
        period = f"{quarter}T{year[-2:]}"
        if period not in periods:
            periods.append(period)
    for year in re.findall(r"\b(20\d{2})\b", text):
        if year not in periods:
            periods.append(year)
    return periods


def markdown_values(text: str) -> list[int | float]:
    cleaned = re.sub(r"<[^>]+>", " ", text)
    cleaned = re.sub(r"!\[[^\]]*\]\([^)]+\)", " ", cleaned)
    cleaned = re.sub(
        r"\(?-?\d{1,3}(?:\.\d{3})*(?:,\d+)?\)?\s*%|\(?-?\d+(?:,\d+)?\)?\s*%",
        " ",
        cleaned,
    )
    tokens = re.findall(r"\(?-?\d{1,3}(?:\.\d{3})*(?:,\d+)?\)?|\(?-?\d+(?:,\d+)?\)?", cleaned)
    values: list[int | float] = []
    for token in tokens:
        value = parse_markdown_number(token)
        if value is not None:
            values.append(value)
    return values


def markdown_value_with_scale(text: str) -> int | float | None:
    cleaned = re.sub(r"<[^>]+>", " ", text)
    match = re.search(
        r"\(?-?\d{1,3}(?:\.\d{3})*(?:,\d+)?\)?|\(?-?\d+(?:,\d+)?\)?",
        cleaned,
    )
    if not match:
        return None
    value = parse_markdown_number(match.group(0))
    if value is None:
        return None
    tail = normalise_text(cleaned[match.end(): match.end() + 40])
    if "milhoes" in tail or "milhao" in tail:
        return value * 1_000_000
    if re.search(r"\bmil\b", tail):
        return value * 1_000
    return value


def markdown_label_text(line: str) -> str:
    cleaned = re.sub(r"<[^>]+>", " ", line)
    cleaned = re.sub(r"[*_`#|]", " ", cleaned)
    cleaned = re.split(r"\(?-?\d", cleaned, maxsplit=1)[0]
    return normalise_text(cleaned)


def markdown_line_text(line: str) -> str:
    cleaned = re.sub(r"<[^>]+>", " ", line)
    cleaned = re.sub(r"[*_`#|]", " ", cleaned)
    return normalise_text(cleaned)


def markdown_table_cells(line: str) -> list[str] | None:
    stripped = line.strip()
    if "|" not in stripped:
        return None
    cells = [re.sub(r"<[^>]+>", " ", cell).strip() for cell in stripped.strip("|").split("|")]
    if len(cells) < 2:
        return None
    if all(re.fullmatch(r":?-{2,}:?", cell.replace(" ", "")) for cell in cells):
        return None
    return cells


def markdown_page_for_line(lines: list[str], index: int) -> int | None:
    for candidate in range(index, max(-1, index - 80), -1):
        text = normalise_text(lines[candidate])
        match = re.search(r"(?:pagina|page)\s*[:#-]?\s*(\d{1,3})", text)
        if match:
            return int(match.group(1))
    return None


def markdown_table_unit(lines: list[str], index: int) -> str | None:
    context = normalise_text(" ".join(lines[max(0, index - 6): index + 1]))
    if "mil pacientes-dia" in context or re.search(r"\bem milhares\b", context):
        return "mil pacientes-dia"
    if "milhares" in context:
        return "milhares"
    if "pacientes-dia" in context:
        return "pacientes-dia"
    return None


def normalize_markdown_table_value(value: int | float, unit: str | None) -> int | float:
    normalized_unit = normalise_text(unit)
    if normalized_unit in {"mil pacientes-dia", "milhares"}:
        return serialisable_number(float(value) * 1_000)
    return value


def extract_metric_from_markdown_tables(company: Company, markdown_paths: list[Path], metric: str) -> list[dict[str, Any]]:
    if company.ticker not in {"MATD3", "RDOR3"} or metric not in {"N. Atendimentos", "N. Pacientes"}:
        return []

    patterns = [re.compile(pattern) for pattern in metric_patterns(company.ticker, metric)]
    results: list[dict[str, Any]] = []
    seen: set[str] = set()

    for path in markdown_paths:
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        if not markdown_matches_company(company, path, text):
            continue

        lines = text.splitlines()
        last_header: tuple[int, list[str], list[str]] | None = None
        for index, line in enumerate(lines):
            cells = markdown_table_cells(line)
            if not cells:
                continue
            periods = [normalise_period(cell) for cell in cells]
            valid_periods = [period for period in periods if period]
            if len(valid_periods) >= 2:
                last_header = (index, cells, [period or "" for period in periods])
                continue
            if last_header is None:
                continue
            label = cells[0]
            full_line = " | ".join(cells)
            if not any(rx.search(normalise_text(label)) or rx.search(normalise_text(full_line)) for rx in patterns):
                continue
            if forbidden_operational_context(company, metric, label, full_line):
                continue
            _header_index, header_cells, period_by_col = last_header
            rdor_markdown_scope = None
            if company.ticker == "RDOR3":
                if not markdown_line_allowed_for_company(company, metric, " ".join(header_cells), lines, index):
                    continue
                rdor_markdown_scope = "Hospitais, oncologia e outros"
            unit = markdown_table_unit(lines, index) or metric_unit(company, metric, label)
            series: dict[str, int | float] = {}
            for col_index, cell in enumerate(cells):
                if col_index >= len(period_by_col):
                    continue
                period = period_by_col[col_index]
                if not period:
                    continue
                value = parse_markdown_number(cell)
                if value is None:
                    continue
                normalized_value = normalize_markdown_table_value(value, unit)
                if plausible_operational_value(label, normalized_value):
                    series[period] = normalized_value
            if not series:
                continue
            series = dict(sorted(series.items(), key=lambda item: period_sort_key(item[0])))
            signature = json.dumps({"metric": metric, "series": series}, ensure_ascii=False, sort_keys=True)
            if signature in seen:
                continue
            seen.add(signature)
            page = markdown_page_for_line(lines, index)
            scope = rdor_markdown_scope or "Release/relatório"
            item = {
                "escopo": scope,
                "fonte_linha": label,
                "fonte_documento": str(path),
                "page": page,
                "unidade": unit,
                "calculado": False,
                "confidence": "high",
                "extraction_method": "release_table",
                "source_type": "release_table",
                "requires_review": False,
                "serie": series,
            }
            enriched = enrich_metric_item(company, metric, item, context=f"{scope} {full_line}")
            if enriched is not None and enriched.get("confidence_score", 0) >= CONFIDENCE_MEDIUM:
                results.append(enriched)
    return sorted(results, key=metric_item_rank)


def markdown_context_is_temporal(context: str) -> bool:
    label = normalise_text(context)
    blocked = (" elimin", "sula", "pro forma", "ajustes")
    return not any(term in label for term in blocked)


def nearest_markdown_period_header(lines: list[str], index: int, lookback: int = 6) -> tuple[str, list[str]]:
    for candidate in range(index, max(-1, index - lookback - 1), -1):
        header = lines[candidate]
        periods = markdown_periods(header)
        if periods:
            return header, periods
    return "", []


def nearby_markdown_context(lines: list[str], index: int, lookback: int = 18) -> str:
    return normalise_text(" ".join(lines[max(0, index - lookback): index + 1]))


def markdown_line_allowed_for_company(
    company: Company,
    metric: str,
    header: str,
    lines: list[str],
    index: int,
) -> bool:
    if not markdown_context_is_temporal(header):
        return False
    if company.ticker != "RDOR3":
        return True

    context = nearby_markdown_context(lines, index)
    forbidden = ("sula", "sulamerica", "sul america", "seguros e previdencia", "consolidado")
    if any(term in context for term in forbidden):
        return False

    # A Rede D'Or deve ser lida pela operação hospitalar/oncológica, sem SulAmérica.
    hospital_terms = ("hospitais", "hospitalar", "oncologia", "rdor")
    if metric in {"Receita Bruta", "Glosa/PCLD", "Ticket Médio", "N. Atendimentos", "N. Unidades", "OcupaÃ§Ã£o"}:
        return any(term in context for term in hospital_terms)
    return True


def markdown_matches_company(company: Company, path: Path, text: str) -> bool:
    haystack = normalise_text(f"{path.stem} {text[:4000]}")
    aliases = COMPANY_MARKDOWN_ALIASES.get(company.ticker, (company.ticker, company.name))
    return any(normalise_text(alias) in haystack for alias in aliases)


def existing_markdown_for_pdf(pdf: Path, markdown_dir: Path) -> Path | None:
    safe_name = normalizar_nome_arquivo(pdf.stem)
    candidates = (
        markdown_dir / f"{safe_name}_markdown" / f"{safe_name}.md",
        markdown_dir / safe_name / f"{safe_name}.md",
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate

    for metadata_path in markdown_dir.rglob("*_metadata.json"):
        try:
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        source_name = Path(str(metadata.get("arquivo_origem", ""))).name
        if source_name and source_name.lower() != pdf.name.lower():
            continue
        markdown = metadata.get("arquivo_markdown")
        if markdown and Path(markdown).exists():
            return Path(markdown)
        normalized_name = metadata.get("nome_documento_normalizado")
        if normalized_name:
            candidate = metadata_path.parent / f"{normalized_name}.md"
            if candidate.exists():
                return candidate
    return None


def ensure_operational_markdowns(markdown_dir: Path, force_parser: bool = False) -> list[Path]:
    markdown_dir.mkdir(parents=True, exist_ok=True)
    if listar_pdfs_entrada is not None:
        for pdf in listar_pdfs_entrada():
            markdown_path = existing_markdown_for_pdf(pdf, markdown_dir)
            if markdown_path and markdown_path.exists() and not force_parser:
                continue
            if converter_pdf_para_markdown is None:
                continue
            converter_pdf_para_markdown(
                pdf,
                diretorio_saida=markdown_dir,
                extrair_imagens=False,
                mostrar_progresso=False,
            )
    return sorted(path for path in markdown_dir.rglob("*.md") if path.is_file())


def extract_metric_from_markdown(company: Company, markdown_paths: list[Path], metric: str) -> list[dict[str, Any]]:
    table_results = extract_metric_from_markdown_tables(company, markdown_paths, metric)
    if table_results:
        return table_results

    patterns = [re.compile(pattern) for pattern in metric_patterns(company.ticker, metric)]
    results: list[dict[str, Any]] = []
    seen: set[str] = set()

    for path in markdown_paths:
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        if not markdown_matches_company(company, path, text):
            continue

        lines = text.splitlines()
        for index, line in enumerate(lines):
            label = markdown_label_text(line)
            full_line = markdown_line_text(line)
            if not any(rx.search(label) or rx.search(full_line) for rx in patterns):
                continue
            normalized_full_line = normalise_text(full_line)
            if metric == "N. Unidades" and (
                "aluguel" in normalized_full_line or "locacao" in normalized_full_line or "r$" in full_line.lower()
            ):
                continue
            if forbidden_operational_context(company, metric, label, full_line):
                continue

            header, periods = nearest_markdown_period_header(lines, index)
            values = markdown_values(line)
            if periods and markdown_line_allowed_for_company(company, metric, header, lines, index) and values:
                useful_values = values[:len(periods)] if len(values) >= len(periods) else values
                useful_periods = periods[-len(useful_values):]
                series = dict(
                    sorted(
                        (
                            (period, value)
                            for period, value in zip(useful_periods, useful_values)
                            if plausible_operational_value(line, value)
                        ),
                        key=lambda item: period_sort_key(item[0]),
                    )
                )
            else:
                nearby = " ".join(lines[max(0, index - 2): index + 3])
                nearby_periods = markdown_periods(nearby)
                value = markdown_value_with_scale(line)
                if not nearby_periods or value is None or not plausible_operational_value(line, value):
                    continue
                period = nearby_periods[-1]
                series = {period: value}

            if not series:
                continue

            signature = json.dumps(series, ensure_ascii=False, sort_keys=True)
            if signature in seen:
                continue
            seen.add(signature)
            context = nearby_markdown_context(lines, index)
            item = {
                "escopo": "Release/relatório",
                "fonte_linha": line.strip()[:180],
                "fonte_documento": str(path),
                "unidade": metric_unit(company, metric, line),
                "calculado": False,
                "confidence": "medium",
                "extraction_method": "markdown_contextual",
                "requires_review": False,
                "serie": series,
            }
            enriched = enrich_metric_item(company, metric, item, context=context)
            if enriched is not None and enriched.get("confidence_score", 0) >= CONFIDENCE_MEDIUM:
                results.append(enriched)
    return sorted(results, key=metric_item_rank)


def enrich_missing_metrics_from_markdown(
    company: Company,
    metrics: dict[str, list[dict[str, Any]]],
    markdown_paths: list[Path],
) -> None:
    for metric, items in metrics.items():
        if items:
            continue
        fallback = extract_metric_from_markdown(company, markdown_paths, metric)
        if fallback:
            metrics[metric] = fallback


def extract_company_from_markdown(
    company: Company,
    markdown_paths: list[Path],
    source_error: str | None = None,
) -> dict[str, Any]:
    metrics: dict[str, list[dict[str, Any]]] = {
        metric: extract_metric_from_markdown(company, markdown_paths, metric)
        for metric in METRIC_NAMES
    }
    prune_trailing_zero_periods(metrics)
    warnings = build_operational_warnings(company, metrics)
    return {
        "ticker": company.ticker,
        "companhia": company.name,
        "fonte_pagina_ri": company.results_url,
        "fonte_planilha": None,
        "arquivo_fundamentos": None,
        "fonte_alternativa": "Markdown gerado pelo app_parser_operacional.py",
        "erro_planilha": source_error,
        "extraido_em_utc": datetime.now(timezone.utc).isoformat(),
        "metricas": metrics,
        "warnings": warnings,
    }


def extract_company(
    company: Company,
    workbook_path: Path,
    source_url: str | None,
    markdown_paths: list[Path] | None = None,
) -> dict[str, Any]:
    raw_workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
        keep_vba=workbook_path.suffix.lower() == ".xlsm",
    )
    # Algumas planilhas de RI trazem dimensões XML infladas (ex.: >1 milhão de
    # linhas por formatação). Os fundamentos reais ficam em poucas centenas de
    # linhas. Criar um snapshot limitado evita percorrer células vazias e deixa
    # a extração previsível sem alterar o arquivo original.
    snapshots: list[SheetSnapshot] = []
    for ws in raw_workbook.worksheets:
        max_row = min(ws.max_row or 2000, 2000)
        max_col = min(ws.max_column or 300, 300)
        rows = [
            tuple(row)
            for row in ws.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_col,
                values_only=True,
            )
        ]
        snapshots.append(SheetSnapshot(ws.title, rows))
    raw_workbook.close()
    workbook = WorkbookSnapshot(snapshots)
    metrics = {metric: extract_metric(company, workbook, metric) for metric in METRIC_NAMES}

    # O Fleury não publica uma linha chamada "Ticket Médio" na planilha anexada.
    # Quando possível, calculamos o indicador a partir de duas linhas divulgadas,
    # mantendo a fórmula explicitamente registrada no JSON.
    if company.ticker == "FLRY3" and not metrics["Ticket Médio"]:
        metrics["Ticket Médio"] = derive_fleury_average_ticket(company, workbook)

    if markdown_paths:
        enrich_missing_metrics_from_markdown(company, metrics, markdown_paths)

    prune_trailing_zero_periods(metrics)
    warnings = build_operational_warnings(company, metrics)

    return {
        "ticker": company.ticker,
        "companhia": company.name,
        "fonte_pagina_ri": company.results_url,
        "fonte_planilha": source_url or str(workbook_path),
        "arquivo_fundamentos": workbook_path.name,
        "extraido_em_utc": datetime.now(timezone.utc).isoformat(),
        "metricas": metrics,
        "warnings": warnings,
    }


def excel_kind(data: bytes) -> str | None:
    """Retorna xlsx/xlsm/xls quando os bytes realmente representam Excel."""
    if data.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
        return ".xls"
    if not data.startswith(b"PK"):
        return None
    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            names = set(archive.namelist())
            if "xl/workbook.xml" not in names:
                return None
            return ".xlsm" if "xl/vbaProject.bin" in names else ".xlsx"
    except zipfile.BadZipFile:
        return None


def candidate_score(text: str, href: str) -> int:
    haystack = normalise_text(f"{text} {unquote(href)}")
    score = 0
    if "planilha de fundamentos" in haystack:
        score += 160
    if "planilha dados historicos" in haystack or "planilha de dados historicos" in haystack:
        score += 150
    if "dados historicos" in haystack:
        score += 80
    if "fundamentos" in haystack:
        score += 80
    if "excel" in haystack:
        score += 40
    if re.search(r"\.(xlsx|xlsm|xls)(?:$|\?)", href, re.IGNORECASE):
        score += 80
    # Em agosto, o último fechamento disponível tende a ser 2T; ainda assim,
    # o score usa apenas o período como desempate e não depende do trimestre atual.
    periods = re.findall(r"([1-4])[tq](\d{2})", haystack)
    if periods:
        q, yy = max((int(q), int(yy)) for q, yy in periods)
        score += yy * 10 + q
    for bad in ("release", "apresentacao", "itr", "dfp", "webcast", "audio", "transcricao"):
        if bad in haystack:
            score -= 120
    return score


async def select_latest_year(frame: Any) -> None:
    """Seleciona o maior ano em combos que efetivamente contenham anos."""
    for select in await frame.locator("select").all():
        try:
            options = await select.locator("option").all()
            years: list[tuple[int, str]] = []
            for option in options:
                text = (await option.inner_text()).strip()
                value = await option.get_attribute("value") or ""
                match = re.search(r"20\d{2}", text)
                if match:
                    years.append((int(match.group()), value))
            if years:
                _year, value = max(years)
                if value:
                    await select.select_option(value=value)
        except Exception:
            # Um select de idioma/navegação não deve interromper o download.
            continue


async def collect_links(page: Any) -> list[tuple[str, str]]:
    found: list[tuple[str, str]] = []
    for frame in page.frames:
        await select_latest_year(frame)
    await page.wait_for_timeout(2500)
    for frame in page.frames:
        try:
            anchors = await frame.locator("a[href]").all()
            for anchor in anchors:
                href = await anchor.get_attribute("href")
                if not href or href.startswith(("javascript:", "mailto:", "tel:", "#")):
                    continue
                text = (await anchor.inner_text()).strip()
                absolute = await anchor.evaluate("a => a.href")
                found.append((text, absolute))
        except Exception:
            continue
    return found


async def download_latest_fundamentals(company: Company, target_dir: Path) -> tuple[Path, str]:
    try:
        from playwright.async_api import async_playwright
    except ImportError as exc:
        raise RuntimeError(
            "Playwright não está instalado. Rode: pip install playwright && playwright install chromium"
        ) from exc

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=True)
        context = await browser.new_context(
            locale="pt-BR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36"
            ),
        )
        page = await context.new_page()
        candidates: dict[str, tuple[int, str]] = {}

        try:
            for page_url in (company.results_url, *company.fallback_pages):
                try:
                    await page.goto(page_url, wait_until="domcontentloaded", timeout=60_000)
                    await page.wait_for_timeout(2500)
                    for text, href in await collect_links(page):
                        score = candidate_score(text, href)
                        if score <= 0:
                            continue
                        previous = candidates.get(href)
                        if previous is None or score > previous[0]:
                            candidates[href] = (score, text)
                except Exception as exc:
                    safe_print(f"[{company.ticker}] aviso ao ler {page_url}: {exc}")

            if not candidates:
                raise RuntimeError("nenhum link candidato a planilha foi encontrado")

            ordered = sorted(candidates.items(), key=lambda item: item[1][0], reverse=True)
            errors: list[str] = []
            for href, (_score, text) in ordered:
                try:
                    response = await context.request.get(href, timeout=60_000, fail_on_status_code=False)
                    if not response.ok:
                        errors.append(f"HTTP {response.status}: {href}")
                        continue
                    data = await response.body()
                    extension = excel_kind(data)
                    if not extension:
                        continue

                    parsed_name = Path(unquote(urlparse(href).path)).name
                    stem = Path(parsed_name).stem if parsed_name else f"{company.ticker}_fundamentos"
                    if not stem or len(stem) > 100:
                        stem = f"{company.ticker}_fundamentos"
                    output = target_dir / f"{stem}{extension}"
                    output.write_bytes(data)
                    safe_print(f"[{company.ticker}] planilha encontrada: {text or href}")
                    return output, href
                except Exception as exc:
                    errors.append(f"{href}: {exc}")

            detail = "; ".join(errors[-3:]) if errors else "candidatos não eram arquivos Excel"
            raise RuntimeError(f"não foi possível baixar uma planilha Excel válida ({detail})")
        finally:
            await browser.close()


async def discover_latest_operational_pdf(ticker: str, target_dir: Path) -> tuple[Path, str]:
    """Discover one recent operational PDF on the company's official RI."""
    from operational_sources import operational_sources_for_sector
    source = operational_sources_for_sector("construcao_civil")[ticker]
    allowed_domain = source["official_domain"].lower()
    try:
        from playwright.async_api import async_playwright
    except ImportError as exc:
        raise RuntimeError("Playwright não está instalado para descoberta de PDFs") from exc
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=True)
        context = await browser.new_context(locale="pt-BR")
        page = await context.new_page()
        candidates: list[tuple[int, str, str]] = []
        try:
            for page_url in source["results_pages"]:
                await page.goto(page_url, wait_until="domcontentloaded", timeout=60000)
                for text, href in await collect_links(page):
                    parsed = urlparse(href)
                    if parsed.netloc.lower() != allowed_domain or not re.search(r"\.pdf(?:$|\?)", parsed.path, re.I):
                        continue
                    haystack = normalise_text(f"{text} {href}")
                    score = 100
                    if any(term in haystack for term in ("release", "resultados", "previa", "apresentacao")): score += 50
                    periods = re.findall(r"([1-4])[tq](\d{2,4})", haystack)
                    if periods: score += max(int(y) for _, y in periods) * 10 + max(int(q) for q, _ in periods)
                    candidates.append((score, href, text))
            if not candidates:
                raise RuntimeError(f"nenhum PDF oficial encontrado para {ticker}")
            _, href, label = max(candidates, key=lambda item: item[0])
            response = await context.request.get(href, timeout=60000, fail_on_status_code=False)
            if not response.ok or not (await response.body()).startswith(b"%PDF"):
                raise RuntimeError(f"PDF oficial inválido ou indisponível: HTTP {response.status}")
            data = await response.body()
            target_dir.mkdir(parents=True, exist_ok=True)
            name = Path(unquote(urlparse(href).path)).name or f"{ticker}_operacional.pdf"
            output = target_dir / f"{ticker}_{name}"
            output.write_bytes(data)
            safe_print(f"[{ticker}] PDF operacional encontrado: {label or href}")
            return output, href
        finally:
            await browser.close()


def write_json(payload: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"{payload['ticker']}.json"
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def collect_operational_observations(payload: dict[str, Any]) -> list[dict[str, Any]]:
    observations: list[dict[str, Any]] = []
    for metric, items in payload.get("metricas", {}).items():
        for item in items:
            for observation in item.get("observations", []):
                row = {
                    "ticker": payload.get("ticker"),
                    "companhia": payload.get("companhia"),
                    **observation,
                    "fonte_pagina_ri": payload.get("fonte_pagina_ri"),
                    "fonte_planilha": payload.get("fonte_planilha"),
                    "arquivo_fundamentos": payload.get("arquivo_fundamentos"),
                }
                observations.append(row)
    return observations


def write_observations_json(observations: list[dict[str, Any]], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / "operational_observations.json"
    payload = {
        "schema_version": "operational_observations_v1",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "observations": observations,
    }
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def parse_local_files(values: list[str]) -> dict[str, Path]:
    parsed: dict[str, Path] = {}
    for value in values:
        if "=" not in value:
            raise ValueError(f"--file deve usar TICKER=caminho; recebido: {value}")
        ticker, raw_path = value.split("=", 1)
        ticker = ticker.strip().upper()
        if ticker not in COMPANIES:
            raise ValueError(f"ticker não suportado: {ticker}")
        path = Path(raw_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(path)
        parsed[ticker] = path
    return parsed


async def run(args: argparse.Namespace) -> int:
    output_dir = Path(args.output_dir).expanduser().resolve()
    if args.sector == "construcao_civil":
        from company_registry import operational_companies
        from construction_operational import CONSTRUCTION_OPERATIONAL_DICTIONARY, calculate_derived_from_observations, extract_markdown_observations, extract_workbook_observations
        from construction_company_profiles import resolve_company_for_document
        from document_catalog import catalog_record, write_catalog
        from tracking import TrackingRun
        markdown_dir = Path(args.md_dir).expanduser().resolve()
        tracker = TrackingRun(sector="construcao_civil", pipeline="operational_extractor", extractor_version="construction_tracking_v1")
        output_dir.mkdir(parents=True, exist_ok=True)
        # Local review PDFs are valid offline fixtures. Convert them to the
        # same Markdown representation used by the production parser.
        if converter_pdf_para_markdown is not None and markdown_dir.exists():
            for review_pdf in markdown_dir.rglob("*.pdf"):
                try:
                    converter_pdf_para_markdown(review_pdf, diretorio_saida=markdown_dir, extrair_imagens=False, mostrar_progresso=False)
                except Exception as exc:
                    safe_print(f"[PDF] aviso: falha ao converter {review_pdf.name}: {exc}")
        allowed_tickers = {company.ticker for company in operational_companies("construcao_civil")}
        # Construction civil is PDF-only. Markdown is accepted only as the
        # local, inspectable text derivative of a PDF; spreadsheets are never
        # an operational source for this sector.
        source_files = [path for pattern in ("*.md", "*.pdf") for path in (markdown_dir.rglob(pattern) if markdown_dir.exists() else ())]
        catalog_records = [catalog_record(path, method="local_official_ri_pdf") for path in source_files if path.suffix.lower() == ".pdf"]
        if not source_files:
            for company in operational_companies("construcao_civil"):
                try:
                    pdf_path, _source_url = await discover_latest_operational_pdf(company.ticker, markdown_dir)
                    if converter_pdf_para_markdown is not None:
                        converter_pdf_para_markdown(pdf_path, diretorio_saida=markdown_dir, extrair_imagens=False, mostrar_progresso=False)
                except Exception as exc:
                    safe_print(f"[{company.ticker}] descoberta de PDF não concluída: {exc}")
            source_files = [path for path in markdown_dir.rglob("*.md") if path.is_file()]
            catalog_records = [catalog_record(path, method="discovered_official_ri_pdf") for path in markdown_dir.rglob("*.pdf") if path.is_file()]
        if catalog_records:
            write_catalog(catalog_records, output_dir / "construction_document_catalog.json")
        all_observations: list[dict[str, Any]] = []
        documents_processed: set[str] = set()
        unresolved_documents: list[dict[str, Any]] = []
        snapshots_preserved = 0
        for company in operational_companies("construcao_civil"):
            observations: list[dict[str, Any]] = []
            company_documents: set[str] = set()
            aliases = tuple(normalise_text(alias) for alias in (company.ticker, *company.legacy_tickers, company.expected_name, *company.aliases))
            for path in source_files:
                document_id = tracker.document(path, source_type="PDF" if path.suffix.lower() == ".pdf" else "Markdown", ticker_hint=path.parent.name if path.parent.name.upper() in allowed_tickers else None)
                tracker.event(document_id, "accepted", source_policy="official_ri_pdf_only")
                path_text = normalise_text(path.name)
                text = ""
                if path.suffix.lower() == ".md":
                    text = path.read_text(encoding="utf-8", errors="replace")
                    tracker.event(document_id, "read", characters=len(text))
                    doc_text = normalise_text(f"{path_text} {text[:2000]}")
                else:
                    # A PDF is a candidate source, but extraction happens via
                    # its Markdown derivative. Do not pass binary PDF bytes to
                    # the textual parser.
                    continue
                resolution = resolve_company_for_document(path.name, text if path.suffix.lower() == ".md" else doc_text, "construcao_civil")
                if not resolution or resolution["ticker"] != company.ticker:
                    if path.suffix.lower() == ".md" and not resolution:
                        unresolved_documents.append({"document": str(path), "reason": "company_unresolved"})
                        tracker.event(document_id, "unresolved", reason="company_unresolved")
                    continue
                company_resolution_method = resolution["method"]
                tracker.event(document_id, "company_resolved", ticker=resolution["ticker"], resolution_method=company_resolution_method, confidence=resolution.get("confidence"))
                # Processamento é contabilizado independentemente de haver
                # observações: ausência de divulgação é um estado auditável.
                company_documents.add(str(path))
                if path.suffix.lower() == ".md":
                    extracted = extract_markdown_observations(text, ticker=company.ticker, source_document=path.name)
                else:
                    extracted = extract_workbook_observations(path, ticker=company.ticker, source_document=path.name)
                    company_resolution_method = "workbook_filename_or_alias"
                for observation in extracted:
                    observation["company_resolution_method"] = company_resolution_method
                    observation["company_resolution"] = resolution
                if extracted:
                    tracker.event(document_id, "parsed", observations_count=len(extracted))
                else:
                    tracker.event(document_id, "validated", observations_count=0, document_without_disclosure=True)
                documents_processed.add(str(path))
                tracker._documents[document_id]["observations_count"] = len(extracted)
                observations.extend(extracted)
            metricas: dict[str, list[dict[str, Any]]] = {}
            for observation in observations:
                name = CONSTRUCTION_OPERATIONAL_DICTIONARY[observation["indicator_id"]]["display_name"]
                metricas.setdefault(name, []).append({"metric": name, "indicator_id": observation["indicator_id"], "confidence": observation["confidence"], "calculated": False, "serie": {observation["period"]: observation["value"]}, "observations": [observation]})
            previous_payload = {}
            previous_path = output_dir / f"{company.ticker}.json"
            if previous_path.exists():
                try:
                    previous_payload = json.loads(previous_path.read_text(encoding="utf-8"))
                except json.JSONDecodeError:
                    previous_payload = {}
            preserved = bool(not observations and previous_payload.get("observations"))
            payload = {
                "schema_version": "construction_operational_v1", "sector": "construcao_civil",
                "generated_at": datetime.now(timezone.utc).isoformat(), "extractor_version": "construction_operational_v1",
                "ticker": company.ticker, "companhia": company.expected_name,
                "companies_requested": len(allowed_tickers), "documents_processed": len(company_documents),
                "metricas": metricas, "observations": observations,
                "status": "found_new_data" if observations else "not_found_no_previous_data",
                "coverage_status": "found" if observations else ("unresolved" if any(item.get("document", "") in {str(p) for p in source_files} for item in unresolved_documents) else "not_found"),
                "discovery": {"source_policy": "official_ri_pdf_only", "documents_processed": sorted(company_documents)},
                "tracking": tracker.summary(),
                "calculation_metadata": calculate_derived_from_observations(observations),
                "warnings": (["Nenhuma métrica operacional válida encontrada nos documentos processados."] if not observations else []) + ["ROE e PCLD/Recebíveis não calculados: dependências financeiras hidratadas não disponíveis nesta extração."],
            }
            if preserved:
                payload = previous_payload
                payload["status"] = "preserved_existing_data"
                payload.setdefault("warnings", []).append("Snapshot anterior preservado: nenhuma observação válida nova encontrada.")
                snapshots_preserved += 1
            write_json(payload, output_dir)
            all_observations.extend(payload.get("observations") or observations)
        observations_by_metric: dict[str, int] = {}
        observations_by_confidence: dict[str, int] = {}
        for observation in all_observations:
            observations_by_metric[observation["indicator_id"]] = observations_by_metric.get(observation["indicator_id"], 0) + 1
            confidence = str(observation.get("confidence") or "unknown")
            observations_by_confidence[confidence] = observations_by_confidence.get(confidence, 0) + 1
        companies_with_observations = sum(1 for path in output_dir.glob("*.json") if path.stem in allowed_tickers and json.loads(path.read_text(encoding="utf-8")).get("observations"))
        result = {
            "sector": "construcao_civil", "companies_requested": len(allowed_tickers),
            "documents_processed": len(documents_processed), "observations_candidates": len(all_observations),
            "observations_valid": len(all_observations), "observations_rejected": 0,
            "snapshot_generated": bool(all_observations),
            "snapshot_path": str(output_dir) if all_observations else None,
            "status": "success_new_snapshot" if all_observations else "no_valid_observations",
            "errors": [] if all_observations else ["Nenhuma observação válida de construcao_civil foi gerada."],
            "warnings": [],
            "unresolved_documents": unresolved_documents,
            "coverage_status": "complete" if companies_with_observations == len(allowed_tickers) else ("partial" if companies_with_observations else "none"),
            "companies_with_documents": len({Path(path).name.split("_", 1)[0] for path in documents_processed}),
            "companies_with_observations": companies_with_observations,
            "companies_without_observations": len(allowed_tickers) - companies_with_observations,
            "operational_files_generated": len(allowed_tickers),
            "snapshots_preserved": snapshots_preserved,
            "observations_by_metric": observations_by_metric,
            "observations_by_confidence": observations_by_confidence,
            "tracking": tracker.summary(),
        }
        if 0 < companies_with_observations < len(allowed_tickers):
            result["status"] = "success_with_warnings"
            result["warnings"] = [f"Cobertura operacional parcial: {companies_with_observations}/{len(allowed_tickers)} empresas com observações."]
        if all_observations:
            write_observations_json(all_observations, output_dir)
        if args.result_json:
            Path(args.result_json).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        tracker.write(output_dir / "tracking" / f"{tracker.run_id}.json")
        safe_print("OPERATIONAL_RESULT=" + json.dumps(result, ensure_ascii=False))
        return 0 if all_observations else 1
    markdown_paths: list[Path] = []
    if not args.no_md_fallback:
        markdown_dir = Path(args.md_dir).expanduser().resolve()
        try:
            markdown_paths = ensure_operational_markdowns(
                markdown_dir,
                force_parser=args.force_md_parser,
            )
            if markdown_paths:
                safe_print(f"[MD] {len(markdown_paths)} documentos Markdown disponíveis para fallback")
        except Exception as exc:
            safe_print(f"[MD] aviso: fallback em Markdown indisponível ({exc})")
    local_files = parse_local_files(args.file)
    tickers = [ticker.upper() for ticker in args.only] if args.only else list(COMPANIES)
    invalid = [ticker for ticker in tickers if ticker not in COMPANIES]
    if invalid:
        raise ValueError(f"tickers não suportados: {', '.join(invalid)}")

    failures: list[tuple[str, str]] = []
    all_observations: list[dict[str, Any]] = []
    with tempfile.TemporaryDirectory(prefix="ri_fundamentos_") as temporary:
        temp_dir = Path(temporary)
        for ticker in tickers:
            company = COMPANIES[ticker]
            try:
                try:
                    if ticker in local_files:
                        workbook_path = local_files[ticker]
                        source_url = None
                    else:
                        workbook_path, source_url = await download_latest_fundamentals(company, temp_dir)

                    payload = extract_company(company, workbook_path, source_url, markdown_paths)
                except Exception as workbook_exc:
                    if not markdown_paths:
                        raise
                    payload = extract_company_from_markdown(
                        company,
                        markdown_paths,
                        source_error=str(workbook_exc),
                    )
                    if not any(payload["metricas"].values()):
                        raise workbook_exc
                    safe_print(f"[{ticker}] aviso: usando Markdown porque a planilha falhou ({workbook_exc})")
                output = write_json(payload, output_dir)
                all_observations.extend(collect_operational_observations(payload))
                found = sum(bool(payload["metricas"][metric]) for metric in METRIC_NAMES)
                safe_print(f"[{ticker}] {output} ({found}/{len(METRIC_NAMES)} indicadores encontrados)")
            except Exception as exc:
                failures.append((ticker, str(exc)))
                safe_print(f"[{ticker}] ERRO: {exc}")

    if failures:
        safe_print("\nFalhas:")
        for ticker, message in failures:
            safe_print(f"  - {ticker}: {message}")
        return 1
    audit_output = write_observations_json(all_observations, output_dir)
    safe_print(f"[AUDITORIA] {audit_output} ({len(all_observations)} observações operacionais)")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sector", choices=("saude", "construcao_civil"), default="saude")
    parser.add_argument("--result-json", default=None)
    parser.add_argument(
        "--output-dir",
        default=".",
        help="diretório dos JSONs (padrão: diretório atual)",
    )
    parser.add_argument(
        "--only",
        nargs="*",
        default=[],
        metavar="TICKER",
        help="processa somente os tickers informados",
    )
    parser.add_argument(
        "--file",
        action="append",
        default=[],
        metavar="TICKER=CAMINHO",
        help="usa uma planilha local em vez de baixar (pode repetir)",
    )
    parser.add_argument(
        "--md-dir",
        default=str(PASTA_MARKDOWNS_OPERACIONAIS),
        help="diretório com os Markdown gerados pelo app_parser_operacional",
    )
    parser.add_argument(
        "--no-md-fallback",
        action="store_true",
        help="desativa a busca em Markdown para indicadores ausentes na planilha",
    )
    parser.add_argument(
        "--force-md-parser",
        action="store_true",
        help="regera os Markdown dos PDFs antes de buscar indicadores ausentes",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    raise SystemExit(asyncio.run(run(args)))


if __name__ == "__main__":
    main()

