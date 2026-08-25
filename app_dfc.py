#!/usr/bin/env python3
"""Baixa ITRs da CVM e monta DFCs históricas em JSON.

Fonte oficial: https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/ITR/DADOS/
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from company_registry import financial_companies
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/ITR/DADOS/"
DFP_BASE_URL = "https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/DFP/DADOS/"
ARQUIVO_RE = re.compile(r"(itr|dfp)_cia_aberta_(\d{4})\.zip", re.I)
DFC_RE = re.compile(r"(itr|dfp)_cia_aberta_DFC_(MD|MI)_(con|ind)_\d{4}\.csv$", re.I)

FATORES_ESCALA = {
    "UNIDADE": 1,
    "UNIDADES": 1,
    "MIL": 1_000,
    "MILHAR": 1_000,
    "MILHARES": 1_000,
    "MILHAO": 1_000_000,
    "MILHOES": 1_000_000,
}


@dataclass(frozen=True)
class Companhia:
    ticker: str
    aliases: tuple[str, ...]
    escopo: str  # con ou ind


COMPANHIAS = (
    Companhia("AALR3", ("ALLIANCA SAUDE E PARTICIPACOES", "CENTRO DE IMAGEM DIAGNOSTICOS", "ALLIAR"), "con"),
    Companhia("DASA3", ("DIAGNOSTICOS DA AMERICA",), "con"),
    Companhia("FLRY3", ("FLEURY",), "con"),
    Companhia("HAPV3", ("HAPVIDA PARTICIPACOES E INVESTIMENTOS", "HAPVIDA"), "con"),
    Companhia("MATD3", ("MATER DEI",), "con"),
    Companhia("ONCO3", ("ONCOCLINICAS DO BRASIL SERVICOS MEDICOS", "ONCOCLINICAS"), "con"),
    Companhia("RDOR3", ("REDE D'OR SAO LUIZ", "REDE DOR SAO LUIZ"), "ind"),
)


def sem_acentos(valor: object) -> str:
    texto = "" if pd.isna(valor) else str(valor)
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", texto.upper()).strip()


def criar_sessao() -> requests.Session:
    sessao = requests.Session()
    sessao.headers.update({"User-Agent": "DFC-CVM-ITR/1.0 (pesquisa financeira)"})
    return sessao


def anos_disponiveis(sessao: requests.Session, doc: str = "itr") -> list[int]:
    base_url = DFP_BASE_URL if doc == "dfp" else BASE_URL
    resposta = sessao.get(base_url, timeout=60)
    resposta.raise_for_status()
    anos = sorted({int(match[1]) for match in ARQUIVO_RE.findall(resposta.text) if match[0].lower() == doc})
    if not anos:
        raise RuntimeError("Não foi possível localizar os ZIPs anuais no índice da CVM.")
    return anos


def zip_valido(caminho: Path, ano: int, doc: str = "itr") -> bool:
    if not caminho.exists() or caminho.stat().st_size == 0:
        return False
    try:
        with zipfile.ZipFile(caminho) as zf:
            membros = [
                nome
                for nome in zf.namelist()
                if DFC_RE.search(Path(nome).name)
                and Path(nome).name.lower().startswith(f"{doc}_cia_aberta")
                and Path(nome).name.endswith(f"_{ano}.csv")
            ]
            return zf.testzip() is None and bool(membros)
    except (OSError, zipfile.BadZipFile):
        return False


def anos_locais(pasta: Path, quantidade: int, doc: str = "itr") -> list[int]:
    anos = []
    for caminho in pasta.glob(f"{doc}_cia_aberta_*.zip"):
        match = ARQUIVO_RE.search(caminho.name)
        if not match:
            continue
        ano = int(match.group(2))
        if zip_valido(caminho, ano, doc):
            anos.append(ano)
    return sorted(set(anos))[-quantidade:]


def baixar(sessao: requests.Session, ano: int, destino: Path, sobrescrever: bool, doc: str = "itr") -> Path:
    destino.mkdir(parents=True, exist_ok=True)
    caminho = destino / f"{doc}_cia_aberta_{ano}.zip"
    if zip_valido(caminho, ano, doc) and not sobrescrever:
        logging.info("ZIP %s ja existe, esta integro e contem CSVs de DFC.", ano)
        return caminho
    if caminho.exists() and not sobrescrever:
        logging.warning("ZIP local %s esta ausente, corrompido ou incompleto; sera baixado novamente.", ano)

    url = f"{DFP_BASE_URL if doc == 'dfp' else BASE_URL}{caminho.name}"
    temporario = caminho.with_suffix(".zip.part")
    logging.info("Baixando %s", url)
    with sessao.get(url, stream=True, timeout=(30, 300)) as resposta:
        resposta.raise_for_status()
        with temporario.open("wb") as arquivo:
            for bloco in resposta.iter_content(chunk_size=1024 * 1024):
                if bloco:
                    arquivo.write(bloco)
    if not zip_valido(temporario, ano, doc):
        raise RuntimeError(f"ZIP de {ano} corrompido ou sem CSVs de DFC esperados.")
    temporario.replace(caminho)
    return caminho


def ler_csv_do_zip(zf: zipfile.ZipFile, membro: str) -> pd.DataFrame:
    with zf.open(membro) as arquivo:
        return pd.read_csv(
            arquivo,
            sep=";",
            encoding="latin1",
            decimal=",",
            dtype={"CNPJ_CIA": "string", "CD_CONTA": "string"},
            low_memory=False,
        )


def localizar_companhia(df: pd.DataFrame, companhia: Companhia) -> tuple[str, str] | None:
    if df.empty or not {"CNPJ_CIA", "DENOM_CIA"}.issubset(df.columns):
        return None
    cadastro = df[["CNPJ_CIA", "DENOM_CIA"]].drop_duplicates().copy()
    cadastro["_nome"] = cadastro["DENOM_CIA"].map(sem_acentos)
    aliases = tuple(sem_acentos(x) for x in companhia.aliases)
    achados = cadastro[cadastro["_nome"].map(lambda n: any(a in n for a in aliases))]
    pares = achados[["CNPJ_CIA", "DENOM_CIA"]].drop_duplicates()
    if len(pares) == 0:
        return None
    if len(pares) > 1:
        raise RuntimeError(
            f"{companhia.ticker}: identificação ambígua. Candidatos: {pares.to_dict('records')}"
        )
    linha = pares.iloc[0]
    return str(linha["CNPJ_CIA"]), str(linha["DENOM_CIA"])


def selecionar_ultima_versao(df: pd.DataFrame) -> pd.DataFrame:
    dados = df.copy()
    dados["DT_REFER"] = pd.to_datetime(dados["DT_REFER"], errors="coerce")
    dados["DT_INI_EXERC"] = pd.to_datetime(dados["DT_INI_EXERC"], errors="coerce")
    dados["DT_FIM_EXERC"] = pd.to_datetime(dados["DT_FIM_EXERC"], errors="coerce")
    dados["VERSAO"] = pd.to_numeric(dados["VERSAO"], errors="coerce")
    def converter_valor(valor: object) -> float:
        texto = str(valor).strip().replace(" ", "")
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        return pd.to_numeric(texto, errors="coerce")

    dados["VL_CONTA_CVM"] = dados["VL_CONTA"].map(converter_valor)
    escalas = dados["ESCALA_MOEDA"].map(sem_acentos)
    escalas_desconhecidas = sorted(set(escalas.dropna()) - set(FATORES_ESCALA))
    if escalas_desconhecidas:
        raise RuntimeError(f"Escala(s) monetaria(s) nao reconhecida(s): {escalas_desconhecidas}")
    dados["FATOR_ESCALA"] = escalas.map(FATORES_ESCALA)
    dados["VL_CONTA"] = dados["VL_CONTA_CVM"] * dados["FATOR_ESCALA"]

    if "ORDEM_EXERC" in dados.columns:
        ordem = dados["ORDEM_EXERC"].map(sem_acentos)
        dados = dados[ordem.str.contains("ULTIMO", na=False)]

    chaves_documento = ["CNPJ_CIA", "DT_REFER"]
    max_versao = dados.groupby(chaves_documento, dropna=False)["VERSAO"].transform("max")
    return dados[dados["VERSAO"].eq(max_versao)].copy()


def extrair(zips: list[Path]) -> tuple[dict[str, list[pd.DataFrame]], list[dict[str, object]]]:
    blocos: dict[str, list[pd.DataFrame]] = {c.ticker: [] for c in COMPANHIAS}
    auditoria: list[dict[str, object]] = []

    for caminho in zips:
        doc_match = ARQUIVO_RE.search(caminho.name)
        documento = doc_match.group(1).upper()
        ano = int(doc_match.group(2))
        with zipfile.ZipFile(caminho) as zf:
            membros = [n for n in zf.namelist() if DFC_RE.search(Path(n).name)]
            if not membros:
                raise RuntimeError(f"Nenhum CSV de DFC encontrado em {caminho.name}.")

            por_escopo: dict[str, list[tuple[str, pd.DataFrame]]] = {"con": [], "ind": []}
            for membro in membros:
                _, metodo, escopo = DFC_RE.search(Path(membro).name).groups()
                por_escopo[escopo.lower()].append((metodo.upper(), ler_csv_do_zip(zf, membro)))

            for companhia in COMPANHIAS:
                encontrados: list[tuple[str, pd.DataFrame, str, str]] = []
                for metodo, df in por_escopo[companhia.escopo]:
                    identidade = localizar_companhia(df, companhia)
                    if identidade is None:
                        continue
                    cnpj, nome = identidade
                    parte = df[df["CNPJ_CIA"].astype(str).eq(cnpj)].copy()
                    if not parte.empty:
                        encontrados.append((metodo, parte, cnpj, nome))

                if not encontrados:
                    logging.warning("%s: DFC %s não localizada em %s; ano mantido como ausente.", companhia.ticker, companhia.escopo, ano)
                    auditoria.append({
                        "ticker": companhia.ticker,
                        "ano_arquivo": ano,
                        "cnpj": None,
                        "denominacao_cvm": None,
                        "escopo": "Consolidado" if companhia.escopo == "con" else "Individual",
                        "metodo_dfc": None,
                        "linhas": 0,
                        "datas_referencia": None,
                        "versoes": None,
                        "status": "DFC ausente no arquivo anual",
                    })
                    continue
                metodos = {x[0] for x in encontrados}
                if len(metodos) > 1:
                    raise RuntimeError(f"{companhia.ticker}/{ano}: há DFC MD e MI; seleção manual necessária.")

                metodo, parte, cnpj, nome = encontrados[0]
                parte = selecionar_ultima_versao(parte)
                parte["METODO_DFC"] = metodo
                parte["ESCOPO"] = "Consolidado" if companhia.escopo == "con" else "Individual"
                parte["TICKER"] = companhia.ticker
                parte["ANO_ARQUIVO"] = ano
                parte["DOCUMENTO_CVM"] = documento
                blocos[companhia.ticker].append(parte)
                auditoria.append({
                    "ticker": companhia.ticker,
                    "ano_arquivo": ano,
                    "cnpj": cnpj,
                    "denominacao_cvm": nome,
                    "escopo": parte["ESCOPO"].iloc[0],
                    "metodo_dfc": metodo,
                    "linhas": len(parte),
                    "datas_referencia": ", ".join(sorted(parte["DT_REFER"].dt.strftime("%Y-%m-%d").dropna().unique())),
                    "versoes": ", ".join(map(str, sorted(parte["VERSAO"].dropna().unique()))),
                    "status": "OK",
                })
    sem_dados = [ticker for ticker, partes in blocos.items() if not partes]
    if sem_dados:
        raise RuntimeError(f"Nenhuma DFC localizada em todos os anos selecionados para: {sem_dados}")
    return blocos, auditoria


def chave_codigo_conta(codigo: object) -> tuple:
    """Ordena códigos CVM hierarquicamente (ex.: 6.01 antes de 6.01.01)."""
    texto = str(codigo)
    return tuple((0, int(x)) if x.isdigit() else (1, x) for x in texto.split("."))


def criar_estrutura_mestre(blocos: dict[str, list[pd.DataFrame]]) -> pd.Index:
    """Cria uma única estrutura de contas, compartilhada por todas as DFCs.

    O código CVM é a chave da estrutura. Quando a descrição de um mesmo código
    varia entre empresas/períodos, adota-se a descrição mais frequente; em
    empate, usa-se a mais recente e, por fim, ordem alfabética para garantir
    resultado determinístico.
    """
    partes = [parte for lista in blocos.values() for parte in lista]
    if not partes:
        raise RuntimeError("Não há dados de DFC para criar a estrutura mestre.")

    dados = pd.concat(partes, ignore_index=True)
    base = dados[["CD_CONTA", "DS_CONTA", "DT_REFER"]].dropna(subset=["CD_CONTA"]).copy()
    base["CD_CONTA"] = base["CD_CONTA"].astype(str)
    base["DS_CONTA"] = base["DS_CONTA"].fillna("").astype(str).str.strip()

    frequencia = (
        base.groupby(["CD_CONTA", "DS_CONTA"], as_index=False)
        .agg(frequencia=("DS_CONTA", "size"), ultima_data=("DT_REFER", "max"))
        .sort_values(
            ["CD_CONTA", "frequencia", "ultima_data", "DS_CONTA"],
            ascending=[True, False, False, True],
        )
    )
    descricao_por_codigo = (
        frequencia.drop_duplicates("CD_CONTA", keep="first")
        .set_index("CD_CONTA")["DS_CONTA"]
        .to_dict()
    )
    codigos = sorted(descricao_por_codigo, key=chave_codigo_conta)
    rotulos = [f"{codigo} | {descricao_por_codigo[codigo]}" for codigo in codigos]
    return pd.Index(rotulos, name="Conta CVM | Descrição")


def preparar_planilha(df: pd.DataFrame, estrutura_mestre: pd.Index) -> pd.DataFrame:
    dados = pd.concat([df], ignore_index=True).drop_duplicates()
    dados["_codigo"] = dados["CD_CONTA"].astype(str)
    dados["_periodo"] = (
        dados["DT_REFER"].dt.strftime("%Y-%m-%d")
        + " | " + dados["DT_INI_EXERC"].dt.strftime("%Y-%m-%d")
        + " a " + dados["DT_FIM_EXERC"].dt.strftime("%Y-%m-%d")
    )
    tabela = dados.pivot_table(index="_codigo", columns="_periodo", values="VL_CONTA", aggfunc="first")

    # Alinha os valores pelo código CVM e depois aplica exatamente a mesma
    # sequência de linhas/rótulos em todas as companhias. Códigos não
    # reportados ficam como NaN, que o Excel grava como célula em branco.
    codigos_mestre = [rotulo.split(" | ", 1)[0] for rotulo in estrutura_mestre]
    tabela = tabela.reindex(codigos_mestre)
    tabela.index = estrutura_mestre
    return tabela


def salvar_excel(blocos: dict[str, list[pd.DataFrame]], auditoria: list[dict[str, object]], saida: Path) -> None:
    saida.parent.mkdir(parents=True, exist_ok=True)
    azul = "1F4E78"
    azul_claro = "D9EAF7"
    estrutura_mestre = criar_estrutura_mestre(blocos)
    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        pd.DataFrame(auditoria).to_excel(writer, sheet_name="Auditoria", index=False)
        for companhia in COMPANHIAS:
            if not blocos[companhia.ticker]:
                continue
            dados = pd.concat(blocos[companhia.ticker], ignore_index=True)
            preparar_planilha(dados, estrutura_mestre).to_excel(writer, sheet_name=companhia.ticker)

        instrucoes = pd.DataFrame({"Informação": [
            "Valores exibidos em reais integrais. VL_CONTA = VL_CONTA_CVM x FATOR_ESCALA.",
            "Cada coluna informa data de referência e intervalo acumulado do fluxo.",
            "Somente ORDEM_EXERC = ÚLTIMO e a maior VERSAO de cada data são usadas.",
            "MD = método direto; MI = método indireto.",
            "Todas as companhias usam exatamente a mesma estrutura de linhas, formada pela união dos códigos de conta CVM encontrados nas sete DFCs.",
            "Quando uma companhia não reporta uma conta da estrutura mestre, a célula correspondente permanece em branco (não é preenchida com zero).",
            "O código CVM é a chave de alinhamento; se a descrição textual variar, uma única descrição canônica é aplicada a todas as abas.",
            "2026 é parcial enquanto nem todos os ITRs do ano tiverem sido entregues.",
        ]})
        instrucoes.to_excel(writer, sheet_name="Leia-me", index=False)

        for ws in writer.book.worksheets:
            ws.freeze_panes = "B2" if ws.title not in ("Auditoria", "Leia-me") else "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=azul)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[1].height = 32
            for col in range(1, ws.max_column + 1):
                largura = 48 if col == 1 else 19
                ws.column_dimensions[get_column_letter(col)].width = largura
            if ws.title not in ("Auditoria", "Leia-me"):
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, 1).fill = PatternFill("solid", fgColor=azul_claro)
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row, col).number_format = '#,##0.00;[Red]-#,##0.00;-'


def valor_json(valor: object) -> object:
    if pd.isna(valor):
        return None
    if isinstance(valor, pd.Timestamp):
        return valor.date().isoformat()
    if hasattr(valor, "item"):
        return valor.item()
    return valor


def registros_json(registros: list[dict[str, object]]) -> list[dict[str, object]]:
    return [
        {chave: valor_json(valor) for chave, valor in registro.items()}
        for registro in registros
    ]


AGREGACOES_DFC = (
    {
        "code": "6.02.AG",
        "description": "Aquisição/alienação de imobilizado - agregado",
        "codes": ("6.02.01", "6.02.02", "6.02.03", "6.02.04"),
    },
    {
        "code": "6.03.CA",
        "description": "Captação de empréstimos e financiamentos - agregado",
        "codes": ("6.03.01", "6.03.04"),
    },
    {
        "code": "6.01.DA",
        "description": "Depreciação e amortização - agregado",
        "codes": ("6.01.01.02", "6.01.01.04"),
    },
)


def adicionar_agregacoes_dfc(linhas: list[dict[str, object]], periodos: list[str]) -> list[dict[str, object]]:
    por_codigo = {str(linha.get("code")): linha for linha in linhas}
    extras = []
    for agregacao in AGREGACOES_DFC:
        valores = {}
        for periodo in periodos:
            soma = 0.0
            encontrou = False
            for codigo in agregacao["codes"]:
                valor = (por_codigo.get(codigo, {}).get("values") or {}).get(periodo)
                if valor is None:
                    continue
                soma += float(valor)
                encontrou = True
            valores[periodo] = soma if encontrou else None
        extras.append(
            {
                "code": agregacao["code"],
                "description": agregacao["description"],
                "depth": str(agregacao["code"]).count("."),
                "synthetic": True,
                "source_codes": list(agregacao["codes"]),
                "values": valores,
            }
        )
        for codigo in agregacao["codes"]:
            if codigo in por_codigo:
                por_codigo[codigo]["parent_code"] = agregacao["code"]

    primeira_fonte = {agregacao["code"]: agregacao["codes"][0] for agregacao in AGREGACOES_DFC}

    def chave_linha(linha: dict[str, object]) -> tuple:
        codigo = str(linha.get("code"))
        if codigo in primeira_fonte:
            return (chave_codigo_conta(primeira_fonte[codigo]), 0, codigo)
        return (chave_codigo_conta(codigo), 1, codigo)

    return sorted(linhas + extras, key=chave_linha)


def montar_empresa_json(ticker: str, dados: pd.DataFrame, estrutura_mestre: pd.Index) -> dict:
    tabela = preparar_planilha(dados, estrutura_mestre).reset_index()
    primeira_coluna = tabela.columns[0]
    periodos = [col for col in tabela.columns if col != primeira_coluna]
    linhas = []
    for linha in tabela.to_dict(orient="records"):
        rotulo = str(linha[primeira_coluna])
        codigo, descricao = rotulo.split(" | ", 1) if " | " in rotulo else (rotulo, "")
        linhas.append(
            {
                "code": codigo,
                "description": descricao,
                "depth": codigo.count("."),
                "values": {periodo: valor_json(linha[periodo]) for periodo in periodos},
            }
        )

    linhas = adicionar_agregacoes_dfc(linhas, periodos)
    metadados = dados.sort_values(["DT_REFER", "VERSAO", "ANO_ARQUIVO"]).iloc[-1]
    return {
        "ticker": ticker,
        "scope": metadados["ESCOPO"],
        "method": metadados["METODO_DFC"],
        "cnpj": metadados["CNPJ_CIA"],
        "denom_cvm": sorted(dados["DENOM_CIA"].dropna().astype(str).unique()),
        "moeda": sorted(dados["MOEDA"].dropna().astype(str).unique()) if "MOEDA" in dados.columns else [],
        "escala_moeda_original": sorted(dados["ESCALA_MOEDA"].dropna().astype(str).unique()) if "ESCALA_MOEDA" in dados.columns else [],
        "unit": "Reais integrais",
        "periods": periodos,
        "rows": linhas,
    }


def salvar_json(blocos: dict[str, list[pd.DataFrame]], auditoria: list[dict[str, object]], saida: Path, anos: list[int]) -> None:
    saida.parent.mkdir(parents=True, exist_ok=True)
    estrutura_mestre = criar_estrutura_mestre(blocos)
    companies = {}
    for companhia in COMPANHIAS:
        if not blocos[companhia.ticker]:
            continue
        dados = pd.concat(blocos[companhia.ticker], ignore_index=True)
        companies[companhia.ticker] = montar_empresa_json(companhia.ticker, dados, estrutura_mestre)

    payload = {
        "kind": "dfc_itr_cvm",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": BASE_URL,
        "years": anos,
        "criteria": "Somente ORDEM_EXERC = ULTIMO e a maior VERSAO de cada data sao usadas.",
        "unit_note": "Valores em reais integrais; VL_CONTA = VL_CONTA_CVM x FATOR_ESCALA.",
        "method_note": "MD = metodo direto; MI = metodo indireto.",
        "companies": companies,
        "audit": registros_json(auditoria),
        "methodology": [
            {"item": "Fonte", "description": BASE_URL},
            {"item": "Anos dos arquivos", "description": ", ".join(map(str, anos))},
            {"item": "Estrutura das DFCs", "description": "Uniao dos codigos de conta CVM encontrados nas companhias, em ordem hierarquica."},
            {"item": "Celulas vazias", "description": "Conta nao reportada pela companhia/periodo; nao representa valor zero."},
        ],
    }
    temporario = saida.with_suffix(".tmp.json")
    temporario.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporario.replace(saida)


def verificar_json(saida: Path) -> None:
    payload = json.loads(saida.read_text(encoding="utf-8"))
    esperado = {c.ticker for c in COMPANHIAS}
    encontrado = set(payload.get("companies", {}))
    faltantes = esperado.difference(encontrado)
    if faltantes:
        raise RuntimeError(f"Empresas ausentes no JSON: {sorted(faltantes)}")
    linhas = {ticker: len(payload["companies"][ticker]["rows"]) for ticker in esperado}
    if len(set(linhas.values())) != 1:
        raise RuntimeError(f"As empresas nao tem a mesma estrutura de linhas: {linhas}")
    if saida.stat().st_size < 1_000:
        raise RuntimeError("O JSON gerado parece incompleto.")


def analisar_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--quantidade-anos", type=int, default=5, help="Anos anuais mais recentes disponíveis (padrão: 5).")
    parser.add_argument("--anos", nargs="+", type=int, help="Anos explícitos; substitui --quantidade-anos.")
    parser.add_argument("--diretorio", type=Path, default=Path("dados_cvm_itr"), help="Pasta de ZIPs e resultados.")
    parser.add_argument("--saida", type=Path, help="Arquivo JSON; padrão: <diretorio>/DFC_ITR_CVM.json.")
    parser.add_argument("--sobrescrever-downloads", action="store_true")
    parser.add_argument("--sem-dfp", action="store_true", help="Nao incorpora DFPs anuais.")
    parser.add_argument("--sector", choices=("saude", "construcao_civil", "all"), default="saude")
    return parser.parse_args()


def main() -> int:
    global COMPANHIAS
    args = analisar_argumentos()
    COMPANHIAS = tuple(Companhia(c.ticker, c.aliases, c.statement_scope) for c in financial_companies(args.sector))
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    sessao = criar_sessao()
    pasta_zips = args.diretorio / "zips"
    try:
        disponiveis = anos_disponiveis(sessao, "itr")
    except requests.RequestException as erro:
        disponiveis = anos_locais(pasta_zips, args.quantidade_anos, "itr")
        if not disponiveis:
            raise
        logging.warning("Nao foi possivel consultar a CVM (%s); usando ZIPs locais: %s", erro, disponiveis)
    anos = sorted(set(args.anos)) if args.anos else disponiveis[-args.quantidade_anos:]
    ausentes = sorted(set(anos) - set(disponiveis))
    if ausentes:
        raise RuntimeError(f"Anos nao disponiveis no portal ou no cache local: {ausentes}")
    logging.info("Anos selecionados: %s", anos)

    zips = [baixar(sessao, ano, pasta_zips, args.sobrescrever_downloads, "itr") for ano in anos]
    if not args.sem_dfp:
        pasta_zips_dfp = args.diretorio / "zips_dfp"
        for ano in anos:
            try:
                zips.append(baixar(sessao, ano, pasta_zips_dfp, args.sobrescrever_downloads, "dfp"))
            except Exception as erro:
                logging.warning("DFP %s nao incorporado (%s).", ano, erro)
    blocos, auditoria = extrair(zips)
    saida = args.saida or args.diretorio / "DFC_ITR_CVM.json"
    salvar_json(blocos, auditoria, saida, anos)
    verificar_json(saida)
    manifesto = args.diretorio / "execucao.json"
    manifesto.write_text(json.dumps({
        "data_execucao": date.today().isoformat(),
        "anos": anos,
        "arquivo_saida": str(saida),
        "companhias": [c.__dict__ for c in COMPANHIAS],
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    logging.info("Concluído: %s", saida.resolve())
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as erro:
        logging.exception("Falha: %s", erro)
        sys.exit(1)
