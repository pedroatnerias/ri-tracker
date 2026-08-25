#!/usr/bin/env python3
"""Baixa ITRs da CVM e gera DREs em Excel para companhias selecionadas.

Uso:
    python cvm_itr_dre.py
    python cvm_itr_dre.py --anos 2022 2023 2024 2025 2026
    python cvm_itr_dre.py --saida resultados/DRE_ITR_CVM.json

Dependencias:
    pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
import io
import json
import logging
import re
import sys
import unicodedata
import urllib.error
import urllib.request
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from company_registry import financial_companies
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PAGINA_CVM = "https://dados.cvm.gov.br/dataset/cia_aberta-doc-itr"
URL_ZIP = "https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/ITR/DADOS/itr_cia_aberta_{ano}.zip"
PAGINA_CVM_DFP = "https://dados.cvm.gov.br/dataset/cia_aberta-doc-dfp"
URL_ZIP_DFP = "https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/DFP/DADOS/dfp_cia_aberta_{ano}.zip"
ARQUIVOS_PADRAO = {
    "consolidado": "itr_cia_aberta_DRE_con_{ano}.csv",
    "individual": "itr_cia_aberta_DRE_ind_{ano}.csv",
}

FATORES_ESCALA = {
    "UNIDADE": 1,
    "UNIDADES": 1,
    "MIL": 1_000,
    "MILHAR": 1_000,
    "MILHARES": 1_000,
    "MILHAO": 1_000_000,
    "MILHOES": 1_000_000,
}


@dataclass(frozen=True)
class Companhia:
    ticker: str
    tipo: str
    # Nomes históricos/atuais normalizados. O primeiro ano encontrado fixa o CD_CVM.
    nomes: tuple[str, ...]


COMPANHIAS = (
    Companhia("AALR3", "consolidado", ("CENTRO DE IMAGEM DIAGNOSTICOS SA", "ALLIANCA SAUDE E PARTICIPACOES SA")),
    Companhia("DASA3", "consolidado", ("DIAGNOSTICOS DA AMERICA SA",)),
    Companhia("FLRY3", "consolidado", ("FLEURY SA",)),
    Companhia("HAPV3", "consolidado", ("HAPVIDA PARTICIPACOES E INVESTIMENTOS SA",)),
    Companhia("MATD3", "consolidado", ("HOSPITAL MATER DEI SA",)),
    Companhia("ONCO3", "consolidado", ("ONCOCLINICAS DO BRASIL SERVICOS MEDICOS SA",)),
    Companhia("RDOR3", "individual", ("REDE DOR SAO LUIZ SA", "REDE DOR SA")),
)

COLUNAS_NECESSARIAS = {
    "CNPJ_CIA", "DT_REFER", "VERSAO", "DENOM_CIA", "CD_CVM", "GRUPO_DFP",
    "MOEDA", "ESCALA_MOEDA", "ORDEM_EXERC", "DT_INI_EXERC", "DT_FIM_EXERC",
    "CD_CONTA", "DS_CONTA", "VL_CONTA", "ST_CONTA_FIXA",
}


def normalizar(texto: object) -> str:
    texto = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    texto = texto.replace("'", "")
    texto = re.sub(r"[^A-Z0-9]+", " ", texto.upper()).strip()
    texto = re.sub(r"\bS A\b", "SA", texto)
    return texto


def requisicao(url: str) -> urllib.request.Request:
    return urllib.request.Request(url, headers={"User-Agent": "CVM-ITR-DRE/1.0 (analise financeira; dados publicos)"})


def descobrir_anos(quantidade: int = 5, doc: str = "itr") -> list[int]:
    """Lê a página do conjunto e devolve os últimos anos com ZIP disponível."""
    pagina_cvm = PAGINA_CVM_DFP if doc == "dfp" else PAGINA_CVM
    prefixo = "dfp" if doc == "dfp" else "itr"
    with urllib.request.urlopen(requisicao(pagina_cvm), timeout=60) as resposta:
        pagina = resposta.read().decode("utf-8", errors="replace")
    anos: set[int] = set()
    anos.update(int(a) for a in re.findall(fr"{prefixo}_cia_aberta_(20\d{{2}})\.zip", pagina, re.I))
    if len(anos) < quantidade:
        # O texto da página também lista os recursos por ano em algumas versões do CKAN.
        anos.update(int(a) for a in re.findall(r"\b(20\d{2})\b", pagina))
    if len(anos) < quantidade:
        raise RuntimeError(f"A página da CVM expôs apenas {len(anos)} ano(s): {sorted(anos)}")
    return sorted(anos)[-quantidade:]


def arquivo_padrao(tipo: str, ano: int, doc: str = "itr") -> str:
    prefixo = "dfp" if doc == "dfp" else "itr"
    return ARQUIVOS_PADRAO[tipo].format(ano=ano).replace("itr_cia_aberta", f"{prefixo}_cia_aberta")


def zip_valido(caminho: Path, ano: int, doc: str = "itr") -> bool:
    if not caminho.exists() or caminho.stat().st_size == 0:
        return False
    esperados = {arquivo_padrao(tipo, ano, doc).lower() for tipo in ARQUIVOS_PADRAO}
    try:
        with zipfile.ZipFile(caminho) as arquivo:
            nomes = {Path(nome).name.lower() for nome in arquivo.namelist()}
            return arquivo.testzip() is None and esperados.issubset(nomes)
    except (OSError, zipfile.BadZipFile):
        return False


def anos_locais(pasta: Path, quantidade: int = 5, doc: str = "itr") -> list[int]:
    anos = []
    prefixo = "dfp" if doc == "dfp" else "itr"
    for caminho in pasta.glob(f"{prefixo}_cia_aberta_*.zip"):
        match = re.search(fr"{prefixo}_cia_aberta_(20\d{{2}})\.zip$", caminho.name, re.I)
        if not match:
            continue
        ano = int(match.group(1))
        if zip_valido(caminho, ano, doc):
            anos.append(ano)
    return sorted(set(anos))[-quantidade:]


def baixar_zip(ano: int, pasta: Path, sobrescrever: bool, doc: str = "itr") -> Path:
    pasta.mkdir(parents=True, exist_ok=True)
    prefixo = "dfp" if doc == "dfp" else "itr"
    destino = pasta / f"{prefixo}_cia_aberta_{ano}.zip"
    if zip_valido(destino, ano, doc) and not sobrescrever:
        logging.info("ZIP %s ja existe, esta integro e contem os CSVs de DRE.", ano)
        return destino
    if destino.exists() and not sobrescrever:
        logging.warning("ZIP local %s esta ausente, corrompido ou incompleto; sera baixado novamente.", ano)

    url = (URL_ZIP_DFP if doc == "dfp" else URL_ZIP).format(ano=ano)
    logging.info("Baixando %s", url)
    temporario = destino.with_suffix(".zip.part")
    with urllib.request.urlopen(requisicao(url), timeout=300) as resposta:
        with temporario.open("wb") as saida:
            while bloco := resposta.read(1024 * 1024):
                if bloco:
                    saida.write(bloco)
    with zipfile.ZipFile(temporario) as arquivo:
        problema = arquivo.testzip()
        if problema:
            raise zipfile.BadZipFile(f"Arquivo interno corrompido: {problema}")
    if not zip_valido(temporario, ano, doc):
        raise zipfile.BadZipFile(f"ZIP de {ano} incompleto ou sem os CSVs de DRE esperados.")
    temporario.replace(destino)
    return destino


def ler_csv_dre(caminho_zip: Path, ano: int, tipo: str, doc: str = "itr") -> pd.DataFrame:
    nome_esperado = arquivo_padrao(tipo, ano, doc)
    with zipfile.ZipFile(caminho_zip) as arquivo:
        nomes = arquivo.namelist()
        candidatos = [n for n in nomes if Path(n).name.lower() == nome_esperado.lower()]
        if not candidatos:
            marcador = f"_DRE_{'con' if tipo == 'consolidado' else 'ind'}_".lower()
            candidatos = [n for n in nomes if marcador in n.lower() and n.lower().endswith(".csv")]
        if len(candidatos) != 1:
            raise RuntimeError(f"Não encontrei univocamente a DRE {tipo} de {ano} em {caminho_zip.name}: {candidatos}")
        dados = arquivo.read(candidatos[0])

    df = pd.read_csv(
        io.BytesIO(dados), sep=";", encoding="ISO-8859-1", dtype=str,
        decimal=",", low_memory=False,
    )
    faltantes = COLUNAS_NECESSARIAS - set(df.columns)
    if faltantes:
        raise RuntimeError(f"Colunas ausentes no leiaute CVM de {ano}: {sorted(faltantes)}")
    df["ANO_ARQUIVO"] = ano
    df["DOCUMENTO_CVM"] = doc.upper()
    return df


def resolver_cd_cvm(base: pd.DataFrame, companhia: Companhia) -> str:
    nomes_alvo = {normalizar(n) for n in companhia.nomes}
    candidatos = base.loc[base["DENOM_CIA"].map(normalizar).isin(nomes_alvo), ["CD_CVM", "DENOM_CIA"]].drop_duplicates()
    codigos = candidatos["CD_CVM"].dropna().unique().tolist()
    if len(codigos) != 1:
        exemplos = candidatos.to_dict("records")
        raise RuntimeError(f"{companhia.ticker}: esperado um único CD_CVM; encontrados {codigos}. Correspondências: {exemplos}")
    return str(codigos[0])


def preparar_dre(base: pd.DataFrame, companhia: Companhia) -> pd.DataFrame:
    cd_cvm = resolver_cd_cvm(base, companhia)
    dre = base.loc[base["CD_CVM"].astype(str) == cd_cvm].copy()
    if dre.empty:
        raise RuntimeError(f"{companhia.ticker}: nenhuma linha após filtrar CD_CVM {cd_cvm}")

    # ÚLTIMO = exercício corrente apresentado no formulário; PENÚLTIMO é comparativo.
    ordem = dre["ORDEM_EXERC"].map(normalizar)
    dre = dre.loc[ordem.str.contains("ULTIMO") & ~ordem.str.contains("PENULTIMO")].copy()
    for coluna in ("DT_REFER", "DT_INI_EXERC", "DT_FIM_EXERC"):
        dre[coluna] = pd.to_datetime(dre[coluna], errors="coerce")
    dre["VERSAO_NUM"] = pd.to_numeric(dre["VERSAO"], errors="coerce").fillna(0)
    def converter_valor(valor: object) -> float:
        texto = str(valor).strip().replace(" ", "")
        # A CVM usa ponto decimal nos arquivos atuais. A segunda forma cobre
        # eventuais leiautes com milhar brasileiro e virgula decimal.
        if "." in texto and "," not in texto:
            return pd.to_numeric(texto, errors="coerce")
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        return pd.to_numeric(texto, errors="coerce")

    # Preserva o número recebido da CVM e converte o valor exibido para reais
    # integrais conforme a escala declarada em cada linha.
    dre["VL_CONTA_CVM"] = dre["VL_CONTA"].map(converter_valor)
    escalas = dre["ESCALA_MOEDA"].map(normalizar)
    escalas_desconhecidas = sorted(set(escalas.dropna()) - set(FATORES_ESCALA))
    if escalas_desconhecidas:
        raise RuntimeError(
            f"{companhia.ticker}: escala(s) monetária(s) não reconhecida(s): "
            f"{escalas_desconhecidas}"
        )
    dre["FATOR_ESCALA"] = escalas.map(FATORES_ESCALA)
    dre["VL_CONTA"] = dre["VL_CONTA_CVM"] * dre["FATOR_ESCALA"]
    if dre[["DT_INI_EXERC", "DT_FIM_EXERC"]].isna().any().any():
        raise RuntimeError(f"{companhia.ticker}: datas inválidas encontradas na DRE")

    # Para cada período/conta, retém a reapresentação mais recente (data de referência e versão).
    chaves = ["CD_CONTA", "DT_INI_EXERC", "DT_FIM_EXERC"]
    dre = dre.sort_values(chaves + ["DT_REFER", "VERSAO_NUM", "ANO_ARQUIVO"])
    dre = dre.drop_duplicates(chaves, keep="last")
    dre["TICKER"] = companhia.ticker
    dre["TIPO_DRE"] = companhia.tipo
    fixas = dre["ST_CONTA_FIXA"].map(normalizar).eq("S")
    dre["CONTA_CHAVE"] = dre["CD_CONTA"].astype(str) + "|" + fixas.map({True: "FIXA", False: "NAO_FIXA"})
    dre.loc[~fixas, "CONTA_CHAVE"] = (
        dre.loc[~fixas, "CONTA_CHAVE"] + "|" + dre.loc[~fixas, "DS_CONTA"].map(normalizar)
    )
    dre["DIAS_PERIODO"] = (dre["DT_FIM_EXERC"] - dre["DT_INI_EXERC"]).dt.days + 1
    dre["PERIODO"] = dre.apply(
        lambda r: f"{r['DT_INI_EXERC']:%d/%m/%Y} a {r['DT_FIM_EXERC']:%d/%m/%Y}", axis=1
    )
    return dre


def chave_conta(codigo: str) -> tuple:
    return tuple(int(p) if p.isdigit() else p for p in re.split(r"[.-]", str(codigo)))


def construir_estrutura_mestra(dres: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Cria uma lista única e ordenada de contas para todas as abas de DRE.

    A estrutura é baseada em CD_CONTA, que é o código hierárquico da CVM. Para
    cada código, usa a descrição mais recente disponível no conjunto completo.
    """
    base = pd.concat(dres.values(), ignore_index=True)
    contas = (
        base.sort_values(["DT_REFER", "VERSAO_NUM", "ANO_ARQUIVO"])
        .drop_duplicates("CONTA_CHAVE", keep="last")
        [["CONTA_CHAVE", "CD_CONTA", "DS_CONTA"]]
        .copy()
    )
    contas["ordem"] = contas["CD_CONTA"].map(chave_conta)
    contas = contas.sort_values(["ordem", "DS_CONTA"]).drop(columns=["ordem"])
    return contas.reset_index(drop=True)


def montar_matriz(dre: pd.DataFrame, estrutura_mestra: pd.DataFrame) -> pd.DataFrame:
    # Reindexar pela estrutura global mantém as mesmas linhas em todas as abas.
    # Contas ausentes na companhia/período permanecem vazias (NaN), não zero.
    periodos = (dre[["PERIODO", "DT_INI_EXERC", "DT_FIM_EXERC"]].drop_duplicates()
                .sort_values(["DT_FIM_EXERC", "DT_INI_EXERC"]))
    pivot = dre.pivot(index="CONTA_CHAVE", columns="PERIODO", values="VL_CONTA")
    pivot = pivot.reindex(columns=periodos["PERIODO"].tolist())
    pivot = pivot.reindex(estrutura_mestra["CONTA_CHAVE"])
    pivot.insert(0, "Descrição da conta", estrutura_mestra["DS_CONTA"].tolist())
    pivot.insert(0, "Código da conta", estrutura_mestra["CD_CONTA"].tolist())
    return pivot.reset_index(drop=True)


def formatar_excel(caminho: Path, tickers: Iterable[str]) -> None:
    wb = load_workbook(caminho)
    azul = "1F4E78"
    azul_claro = "D9EAF7"
    for nome in [*tickers, "Base_Auditoria", "Metodologia"]:
        ws = wb[nome]
        ws.freeze_panes = "C2" if nome in tickers else "A2"
        ws.auto_filter.ref = ws.dimensions
        for celula in ws[1]:
            celula.fill = PatternFill("solid", fgColor=azul)
            celula.font = Font(color="FFFFFF", bold=True)
            celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32
        if nome in tickers:
            ws.column_dimensions["A"].width = 17
            ws.column_dimensions["B"].width = 55
            for coluna in range(3, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(coluna)].width = 24
                for linha in range(2, ws.max_row + 1):
                    ws.cell(linha, coluna).number_format = '#,##0;[Red]-#,##0;–'
            for linha in range(2, ws.max_row + 1):
                codigo = str(ws.cell(linha, 1).value or "")
                nivel = max(0, codigo.count(".") - 1)
                ws.cell(linha, 2).alignment = Alignment(indent=min(nivel, 7))
                if codigo.count(".") <= 1:
                    for celula in ws[linha]:
                        celula.font = Font(bold=True)
                        celula.fill = PatternFill("solid", fgColor=azul_claro)
        else:
            for coluna in ws.columns:
                letra = get_column_letter(coluna[0].column)
                largura = min(45, max(12, max(len(str(c.value or "")) for c in coluna[:200]) + 2))
                ws.column_dimensions[letra].width = largura
    wb.save(caminho)


def exportar_excel(dres: dict[str, pd.DataFrame], caminho: Path, anos: list[int]) -> None:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    auditoria = pd.concat(dres.values(), ignore_index=True)
    estrutura_mestra = construir_estrutura_mestra(dres)
    colunas_auditoria = [
        "TICKER", "TIPO_DRE", "CNPJ_CIA", "CD_CVM", "DENOM_CIA", "GRUPO_DFP",
        "MOEDA", "ESCALA_MOEDA", "FATOR_ESCALA", "DT_REFER", "VERSAO", "DT_INI_EXERC",
        "DT_FIM_EXERC", "DIAS_PERIODO", "CD_CONTA", "DS_CONTA", "VL_CONTA_CVM", "VL_CONTA",
        "ANO_ARQUIVO",
    ]
    metodologia = pd.DataFrame({"Item": [
        "Fonte", "Anos dos arquivos", "Escopo", "Tipo por companhia", "Reapresentações",
        "Períodos", "Estrutura das DREs", "Células vazias", "Unidade monetária", "Limitação do ITR",
    ], "Descrição": [
        PAGINA_CVM,
        ", ".join(map(str, anos)),
        "Demonstração do Resultado (DRE), linhas fixas e não fixas da CVM.",
        "; ".join(f"{c.ticker}: {c.tipo}" for c in COMPANHIAS),
        "Para a mesma conta e período, foi mantida a linha de maior DT_REFER e VERSAO.",
        "Somente ORDEM_EXERC = ÚLTIMO; datas de início e fim são as informadas pela companhia.",
        "Todas as abas usam a mesma lista mestra, formada pela união dos códigos de conta encontrados nas sete companhias, em ordem hierárquica da CVM.",
        "Uma célula vazia indica que a conta não foi divulgada para aquela companhia/período; não representa valor zero.",
        "Valores exibidos em reais integrais, sem abreviação. VL_CONTA = VL_CONTA_CVM × FATOR_ESCALA; a escala original permanece em ESCALA_MOEDA.",
        "ITR cobre informações trimestrais; o 4º trimestre/ano completo normalmente é divulgado no DFP, não neste conjunto.",
    ]})
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        for ticker, dre in dres.items():
            montar_matriz(dre, estrutura_mestra).to_excel(writer, sheet_name=ticker, index=False)
        auditoria[colunas_auditoria].to_excel(writer, sheet_name="Base_Auditoria", index=False)
        metodologia.to_excel(writer, sheet_name="Metodologia", index=False)
    formatar_excel(caminho, dres.keys())


def valor_json(valor: object) -> object:
    if pd.isna(valor):
        return None
    if isinstance(valor, pd.Timestamp):
        return valor.date().isoformat()
    if hasattr(valor, "item"):
        return valor.item()
    return valor


def dataframe_json(df: pd.DataFrame) -> list[dict]:
    registros = []
    for registro in df.to_dict(orient="records"):
        registros.append({chave: valor_json(valor) for chave, valor in registro.items()})
    return registros


def montar_empresa_json(ticker: str, dre: pd.DataFrame, estrutura_mestra: pd.Series) -> dict:
    matriz = montar_matriz(dre, estrutura_mestra)
    periodos = [col for col in matriz.columns if col not in ("Código da conta", "Descrição da conta")]
    periodos_base = (
        dre[["PERIODO", "DT_INI_EXERC", "DT_FIM_EXERC"]]
        .drop_duplicates()
        .set_index("PERIODO")
    )
    period_metadata = {}
    for periodo in periodos:
        if periodo not in periodos_base.index:
            continue
        row = periodos_base.loc[periodo]
        inicio = pd.Timestamp(row["DT_INI_EXERC"])
        fim = pd.Timestamp(row["DT_FIM_EXERC"])
        period_metadata[periodo] = {
            "start_date": inicio.date().isoformat(),
            "end_date": fim.date().isoformat(),
            "year": int(fim.year),
            "quarter": int((fim.month - 1) // 3 + 1),
            "is_ytd": bool(inicio.month == 1 and inicio.day == 1),
        }

    by_year_quarter = {
        (meta["year"], meta["quarter"], meta["is_ytd"]): periodo
        for periodo, meta in period_metadata.items()
    }
    for year in sorted({meta["year"] for meta in period_metadata.values()}):
        annual = by_year_quarter.get((year, 4, True))
        q1 = by_year_quarter.get((year, 1, True))
        q2 = by_year_quarter.get((year, 2, False))
        q3 = by_year_quarter.get((year, 3, False))
        if not all((annual, q1, q2, q3)):
            continue
        q4 = f"01/10/{year} a 31/12/{year}"
        if q4 in matriz.columns:
            continue
        matriz[q4] = matriz.apply(
            lambda row: (
                row[annual] - row[q1] - row[q2] - row[q3]
                if pd.notna(row[annual]) and pd.notna(row[q1]) and pd.notna(row[q2]) and pd.notna(row[q3])
                else None
            ),
            axis=1,
        )
        period_metadata[q4] = {
            "start_date": f"{year}-10-01",
            "end_date": f"{year}-12-31",
            "year": int(year),
            "quarter": 4,
            "is_ytd": False,
            "derived": "DFP anual - (1T + 2T + 3T)",
        }
        periodos.append(q4)

    periodos = sorted(
        periodos,
        key=lambda periodo: (
            period_metadata.get(periodo, {}).get("end_date", periodo),
            period_metadata.get(periodo, {}).get("start_date", periodo),
        ),
    )
    linhas = []
    for row in matriz.to_dict(orient="records"):
        codigo = str(row["Código da conta"])
        linhas.append(
            {
                "code": codigo,
                "description": row["Descrição da conta"],
                "depth": codigo.count("."),
                "values": {periodo: valor_json(row[periodo]) for periodo in periodos},
            }
        )

    metadados = dre.sort_values(["DT_REFER", "VERSAO_NUM", "ANO_ARQUIVO"]).iloc[-1]
    return {
        "ticker": ticker,
        "tipo_dre": metadados["TIPO_DRE"],
        "cd_cvm": str(metadados["CD_CVM"]),
        "cnpj": metadados["CNPJ_CIA"],
        "denom_cvm": sorted(dre["DENOM_CIA"].dropna().astype(str).unique()),
        "moeda": sorted(dre["MOEDA"].dropna().astype(str).unique()),
        "escala_moeda_original": sorted(dre["ESCALA_MOEDA"].dropna().astype(str).unique()),
        "unit": "Reais integrais",
        "periods": periodos,
        "period_metadata": period_metadata,
        "rows": linhas,
    }


def exportar_json(dres: dict[str, pd.DataFrame], caminho: Path, anos: list[int]) -> None:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    auditoria = pd.concat(dres.values(), ignore_index=True)
    estrutura_mestra = construir_estrutura_mestra(dres)
    colunas_auditoria = [
        "TICKER", "TIPO_DRE", "CNPJ_CIA", "CD_CVM", "DENOM_CIA", "GRUPO_DFP",
        "MOEDA", "ESCALA_MOEDA", "FATOR_ESCALA", "DT_REFER", "VERSAO", "DT_INI_EXERC",
        "DT_FIM_EXERC", "DIAS_PERIODO", "CD_CONTA", "DS_CONTA", "VL_CONTA_CVM", "VL_CONTA",
        "ANO_ARQUIVO",
    ]
    payload = {
        "kind": "dre_itr_cvm",
        "generated_at": pd.Timestamp.now().isoformat(timespec="seconds"),
        "source": PAGINA_CVM,
        "years": anos,
        "criteria": "Para a mesma conta e periodo, foi mantida a linha de maior DT_REFER e VERSAO; apenas ORDEM_EXERC = ULTIMO.",
        "unit_note": "Valores em reais integrais; VL_CONTA = VL_CONTA_CVM x FATOR_ESCALA.",
        "companies": {
            ticker: montar_empresa_json(ticker, dre, estrutura_mestra)
            for ticker, dre in dres.items()
        },
        "audit": dataframe_json(auditoria[colunas_auditoria]),
        "methodology": [
            {"item": "Fonte", "description": PAGINA_CVM},
            {"item": "Anos dos arquivos", "description": ", ".join(map(str, anos))},
            {"item": "Escopo", "description": "Demonstracao do Resultado (DRE), linhas fixas e nao fixas da CVM."},
            {"item": "Tipo por companhia", "description": "; ".join(f"{c.ticker}: {c.tipo}" for c in COMPANHIAS)},
            {"item": "Celulas vazias", "description": "Conta nao divulgada para a companhia/periodo; nao representa valor zero."},
        ],
    }
    temporario = caminho.with_suffix(".tmp.json")
    temporario.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporario.replace(caminho)


def verificar_json(caminho: Path, dres: dict[str, pd.DataFrame]) -> None:
    payload = json.loads(caminho.read_text(encoding="utf-8"))
    esperado = set(dres)
    encontrado = set(payload.get("companies", {}))
    faltantes = esperado.difference(encontrado)
    if faltantes:
        raise RuntimeError(f"Empresas ausentes no JSON: {sorted(faltantes)}")
    linhas = {ticker: len(payload["companies"][ticker]["rows"]) for ticker in esperado}
    if len(set(linhas.values())) != 1:
        raise RuntimeError(f"As empresas nao tem a mesma estrutura de linhas: {linhas}")
    if caminho.stat().st_size < 1_000:
        raise RuntimeError("O JSON gerado parece incompleto.")


def analisar_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Baixa ITRs e monta DREs no formato de contas da CVM.")
    parser.add_argument("--anos", nargs="+", type=int, help="Anos específicos; por padrão, os cinco mais recentes da página.")
    parser.add_argument("--pasta-zips", type=Path, default=Path("dados_cvm") / "zips")
    parser.add_argument("--pasta-zips-dfp", type=Path, default=Path("dados_cvm") / "zips_dfp")
    parser.add_argument("--saida", type=Path, default=Path("resultados") / "DRE_ITR_CVM_ultimos_5_anos.json")
    parser.add_argument("--sobrescrever-zips", action="store_true")
    parser.add_argument("--sem-dfp", action="store_true", help="Nao incorpora DFPs anuais.")
    parser.add_argument("--sector", choices=("saude", "construcao_civil", "all"), default="saude")
    return parser.parse_args()


def main() -> int:
    global COMPANHIAS
    args = analisar_argumentos()
    COMPANHIAS = tuple(Companhia(c.ticker, "consolidado" if c.statement_scope == "con" else "individual", c.aliases) for c in financial_companies(args.sector))
    logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
    if args.anos:
        anos = sorted(set(args.anos))
    else:
        try:
            anos = descobrir_anos(5, "itr")
        except urllib.error.URLError as erro:
            anos = anos_locais(args.pasta_zips, 5, "itr")
            if not anos:
                raise
            logging.warning("Nao foi possivel consultar a CVM (%s); usando ZIPs locais: %s", erro, anos)
    logging.info("Anos selecionados: %s", anos)

    por_tipo: dict[str, list[pd.DataFrame]] = {"consolidado": [], "individual": []}
    for ano in anos:
        caminho_zip = baixar_zip(ano, args.pasta_zips, args.sobrescrever_zips, "itr")
        for tipo in por_tipo:
            por_tipo[tipo].append(ler_csv_dre(caminho_zip, ano, tipo, "itr"))
    if not args.sem_dfp:
        for ano in anos:
            try:
                caminho_zip = baixar_zip(ano, args.pasta_zips_dfp, args.sobrescrever_zips, "dfp")
            except Exception as erro:
                logging.warning("DFP %s nao incorporado (%s).", ano, erro)
                continue
            for tipo in por_tipo:
                por_tipo[tipo].append(ler_csv_dre(caminho_zip, ano, tipo, "dfp"))
    bases = {tipo: pd.concat(partes, ignore_index=True) for tipo, partes in por_tipo.items()}

    dres: dict[str, pd.DataFrame] = {}
    for companhia in COMPANHIAS:
        dres[companhia.ticker] = preparar_dre(bases[companhia.tipo], companhia)
        logging.info("%s: %d linhas selecionadas.", companhia.ticker, len(dres[companhia.ticker]))
    exportar_json(dres, args.saida, anos)
    verificar_json(args.saida, dres)
    logging.info("Concluído: %s", args.saida.resolve())
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (urllib.error.URLError, RuntimeError, zipfile.BadZipFile, ValueError) as erro:
        logging.error("%s", erro)
        raise SystemExit(1)
